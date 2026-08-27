"""
create_alias_templates.py
--------------------------
Creates a blank Excel template for company/org name aliases.

Output file:
  org_aliases_template.xlsx — company name aliases

Two columns:
  alias      — the raw variant as it appears in CT.gov (e.g. "Pfizer Inc.")
  canonical  — your preferred display name (e.g. "Pfizer")

After filling in your data, load with:
  python manage_aliases.py import org org_aliases_template.xlsx

Notes:
  Drug aliases    → managed in dept_keywords (drug_name / alias_names)
  Condition aliases → managed in dept_indications (indication / keywords)
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent

CONFIGS = [
    {
        "filename": "org_aliases_template.xlsx",
        "title": "Organization Aliases",
        "alias_header": "alias  (raw CT.gov name)",
        "canonical_header": "canonical  (preferred display name)",
        "examples": [
            ("Pfizer Inc.", "Pfizer"),
            ("Merck Sharp & Dohme LLC", "Merck Sharp & Dohme"),
            ("F. Hoffmann-La Roche Ltd", "F. Hoffmann-La Roche"),
        ],
    },
]

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="EFF6FF")
EXAMPLE_FONT = Font(color="374151", name="Calibri", size=10, italic=True)


def make_template(config: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Aliases"

    # Column widths
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 45

    # Header row
    ws["A1"] = "alias"
    ws["B1"] = "canonical"
    for col in ("A1", "B1"):
        ws[col].font      = HEADER_FONT
        ws[col].fill      = HEADER_FILL
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Example rows (greyed out)
    for i, (alias, canonical) in enumerate(config["examples"], start=2):
        ws[f"A{i}"] = alias
        ws[f"B{i}"] = canonical
        ws[f"A{i}"].font = EXAMPLE_FONT
        ws[f"B{i}"].font = EXAMPLE_FONT
        ws[f"A{i}"].fill = EXAMPLE_FILL
        ws[f"B{i}"].fill = EXAMPLE_FILL

    out = HERE / config["filename"]
    wb.save(out)
    print(f"  Created: {out}")


def main():
    print("Creating alias template files...")
    for cfg in CONFIGS:
        make_template(cfg)
    print("\nDone.")
    print("Fill in your company aliases and import with:")
    print("  python manage_aliases.py import org org_aliases_template.xlsx")


if __name__ == "__main__":
    main()
