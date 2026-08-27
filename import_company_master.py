"""
import_company_master.py
------------------------
Imports the ClinicalTrials Company Normalization Master Excel into the DB.

Reads three sheets:
  - "Company Master"  -> org_master (canonical names + metadata)
                      -> org_aliases (alias variants -> canonical)
  - "Do Not Merge"    -> validation check: ensures no two "do not merge" companies
                        share an alias that would accidentally collapse them
  - "Methodology"     -> informational only, not imported

Tables affected:
  - org_master  (created if not exists): canonical_name, entity_type,
                relationship, parent_company
  - org_aliases (upserted): alias -> canonical_name

Usage:
    python import_company_master.py
    python import_company_master.py --dry-run   # preview only, no DB writes
"""

import sys
import openpyxl
import psycopg2.extras
import db

EXCEL_PATH = r"C:\Users\LAPTOP\Downloads\ClinicalTrials_Company_Normalization_Master.xlsx"
DRY_RUN    = "--dry-run" in sys.argv

# ── DDL ───────────────────────────────────────────────────────────────────────

_CREATE_ORG_MASTER = """
CREATE TABLE IF NOT EXISTS org_master (
    canonical_name  TEXT PRIMARY KEY,
    entity_type     TEXT,
    relationship    TEXT,
    parent_company  TEXT,
    updated_at      TIMESTAMPTZ DEFAULT NOW()
);
"""

# ── Helpers ───────────────────────────────────────────────────────────────────

def _cell(value) -> str:
    """Safely convert a cell value to a stripped string, empty string if None."""
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("none", "nan", "-") else s


def _split_aliases(raw: str) -> list[str]:
    """Split a pipe-separated alias string into a cleaned list."""
    if not raw:
        return []
    return [a.strip() for a in raw.split("|") if a.strip()]


# ── Read Excel ────────────────────────────────────────────────────────────────

def read_company_master(wb: openpyxl.Workbook) -> tuple[list[dict], list[tuple[str, str]]]:
    """
    Returns:
      master_rows : list of dicts for org_master
      alias_pairs : list of (alias, canonical_name) for org_aliases
    """
    ws = wb["Company Master"]
    rows = list(ws.iter_rows(values_only=True))
    header = [_cell(c).lower() for c in rows[0]]

    # Locate columns flexibly
    def _col(name):
        for i, h in enumerate(header):
            if name in h:
                return i
        return None

    ci_name     = _col("company name")
    ci_alias    = _col("alias")
    ci_type     = _col("entity type")
    ci_rel      = _col("relationship")
    ci_parent   = _col("parent")

    if ci_name is None:
        print("ERROR: 'Company Name' column not found in Company Master sheet.")
        sys.exit(1)

    master_rows : list[dict]         = []
    alias_pairs : list[tuple[str, str]] = []
    skipped = 0

    for row in rows[1:]:
        canonical = _cell(row[ci_name]) if ci_name is not None else ""
        if not canonical:
            skipped += 1
            continue

        entity_type    = _cell(row[ci_type])   if ci_type   is not None else ""
        relationship   = _cell(row[ci_rel])    if ci_rel    is not None else ""
        parent_company = _cell(row[ci_parent]) if ci_parent is not None else ""
        alias_raw      = _cell(row[ci_alias])  if ci_alias  is not None else ""

        master_rows.append({
            "canonical_name": canonical,
            "entity_type":    entity_type or None,
            "relationship":   relationship or None,
            "parent_company": parent_company or None,
        })

        # The canonical name itself maps to itself (catches exact CT.gov matches)
        alias_pairs.append((canonical, canonical))

        # Each pipe-separated alias variant
        for alias in _split_aliases(alias_raw):
            if alias.lower() != canonical.lower():
                alias_pairs.append((alias, canonical))

    return master_rows, alias_pairs


def read_do_not_merge(wb: openpyxl.Workbook) -> list[tuple[str, str, str]]:
    """Returns list of (company_a, company_b, reason) from Do Not Merge sheet."""
    if "Do Not Merge" not in wb.sheetnames:
        return []
    ws  = wb["Do Not Merge"]
    rows = list(ws.iter_rows(values_only=True))
    result = []
    for row in rows[1:]:
        a      = _cell(row[0]) if len(row) > 0 else ""
        b      = _cell(row[1]) if len(row) > 1 else ""
        reason = _cell(row[2]) if len(row) > 2 else ""
        if a and b:
            result.append((a, b, reason))
    return result


# ── Validation ────────────────────────────────────────────────────────────────

def validate_do_not_merge(
    alias_pairs: list[tuple[str, str]],
    do_not_merge: list[tuple[str, str, str]],
) -> list[str]:
    """
    Check that no alias accidentally maps two "Do Not Merge" companies
    to the same canonical name.
    Returns a list of warning strings (empty = all clear).
    """
    # Build alias -> canonical lookup
    alias_map: dict[str, str] = {}
    for alias, canonical in alias_pairs:
        alias_map[alias.lower()] = canonical

    warnings = []
    for a, b, reason in do_not_merge:
        canon_a = alias_map.get(a.lower(), a)
        canon_b = alias_map.get(b.lower(), b)
        if canon_a.lower() == canon_b.lower():
            warnings.append(
                f"  CONFLICT: '{a}' and '{b}' both resolve to '{canon_a}'\n"
                f"    Reason: {reason}"
            )
    return warnings


# ── DB writes ─────────────────────────────────────────────────────────────────

def upsert_org_master(conn, rows: list[dict]) -> int:
    sql = """
        INSERT INTO org_master (canonical_name, entity_type, relationship, parent_company, updated_at)
        VALUES %s
        ON CONFLICT (canonical_name)
        DO UPDATE SET
            entity_type    = EXCLUDED.entity_type,
            relationship   = EXCLUDED.relationship,
            parent_company = EXCLUDED.parent_company,
            updated_at     = NOW()
    """
    data = [(r["canonical_name"], r["entity_type"], r["relationship"], r["parent_company"]) for r in rows]
    # Append NOW() placeholder handled via psycopg2 — use template
    sql_with_ts = """
        INSERT INTO org_master (canonical_name, entity_type, relationship, parent_company)
        VALUES %s
        ON CONFLICT (canonical_name)
        DO UPDATE SET
            entity_type    = EXCLUDED.entity_type,
            relationship   = EXCLUDED.relationship,
            parent_company = EXCLUDED.parent_company,
            updated_at     = NOW()
    """
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(cur, sql_with_ts, data, page_size=200)
    return len(rows)


def upsert_org_aliases(conn, pairs: list[tuple[str, str]]) -> int:
    # Deduplicate by alias key — last definition wins within the batch
    deduped: dict[str, str] = {}
    for alias, canonical in pairs:
        deduped[alias.lower()] = (alias, canonical)
    unique_pairs = list(deduped.values())

    sql = """
        INSERT INTO org_aliases (alias, canonical)
        VALUES %s
        ON CONFLICT (alias)
        DO UPDATE SET canonical = EXCLUDED.canonical
    """
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(cur, sql, unique_pairs, page_size=500)
    return len(unique_pairs)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print(f"Import Company Master{' [DRY RUN]' if DRY_RUN else ''}")
    print("=" * 60)

    # ── Read Excel ────────────────────────────────────────────────
    print("\nReading Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    master_rows, alias_pairs = read_company_master(wb)
    do_not_merge = read_do_not_merge(wb)
    wb.close()

    print(f"  Companies (org_master rows) : {len(master_rows)}")
    print(f"  Alias pairs (org_aliases)   : {len(alias_pairs)}")
    print(f"  Do Not Merge rules          : {len(do_not_merge)}")

    # ── Validate Do Not Merge ─────────────────────────────────────
    print("\nValidating Do Not Merge rules...")
    warnings = validate_do_not_merge(alias_pairs, do_not_merge)
    if warnings:
        print(f"  WARNING: {len(warnings)} conflict(s) found — review before importing:")
        for w in warnings:
            print(w)
        if not DRY_RUN:
            ans = input("\nConflicts found. Continue anyway? (y/N): ").strip().lower()
            if ans != "y":
                print("Aborted.")
                return
    else:
        print("  OK: No conflicts — all Do Not Merge rules are satisfied.")

    # ── Preview (dry run) ─────────────────────────────────────────
    if DRY_RUN:
        print("\n[DRY RUN] Sample alias pairs:")
        for alias, canonical in alias_pairs[:15]:
            if alias != canonical:
                print(f"  '{alias}' -> '{canonical}'")
        print(f"\n[DRY RUN] Sample org_master rows:")
        for r in master_rows[:10]:
            print(f"  {r['canonical_name']!r:40}  {r['entity_type'] or '—':25}  {r['parent_company'] or ''}")
        print("\n[DRY RUN] No changes written.")
        return

    # ── DB writes ─────────────────────────────────────────────────
    conn = db._db()

    print("\nCreating org_master table (if not exists)...")
    with conn.cursor() as cur:
        cur.execute(_CREATE_ORG_MASTER)
    print("  Done.")

    print("Upserting org_master...")
    n_master = upsert_org_master(conn, master_rows)
    print(f"  {n_master} rows upserted.")

    print("Upserting org_aliases...")
    n_aliases = upsert_org_aliases(conn, alias_pairs)
    print(f"  {n_aliases} alias pairs upserted.")

    # ── Summary ───────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("Import complete.")
    print(f"  org_master  : {n_master} companies")
    print(f"  org_aliases : {n_aliases} aliases")
    if warnings:
        print(f"  Warnings    : {len(warnings)} Do Not Merge conflicts (review above)")
    print("\nNext: run backfill_normalization.py to apply new aliases to existing data.")
    print("=" * 60)


if __name__ == "__main__":
    main()
