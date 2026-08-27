"""
migrate_to_db.py — One-time migration from Excel/JSON files to Supabase.

Run ONCE before switching the pipeline to use the database:
    python migrate_to_db.py ADC

What it migrates:
    ADC_All_NCTIDs.xlsx          → tracking_list table
    state.json                   → pipeline_state table
    latest_comparison.xlsx       → version_cache table
    Clinical_Trials_Organized.xlsx → organized_trials table
    ADC/output/combined_*.xlsx   → new_candidates_log, unmatched_log, run_history
"""

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, r"C:\Users\LAPTOP\Downloads\NCT_Record_History_Changes")

import db
from NCT_Changes_Tracker import load_nct_ids_from_excel


def migrate_tracking_list(dept: str, tracking_xlsx: Path):
    print("\n[1] Migrating tracking list...")
    nct_ids = load_nct_ids_from_excel(str(tracking_xlsx))
    db.upsert_tracking_list(dept, nct_ids, added_by="migration")
    print(f"  Migrated {len(nct_ids)} NCT IDs")


def migrate_pipeline_state(dept: str, state_json: Path):
    print("\n[2] Migrating pipeline state...")
    if not state_json.exists():
        print("  No state.json found — skipping (will start fresh)")
        return
    state = json.loads(state_json.read_text(encoding="utf-8"))
    db.save_pipeline_state(
        dept=dept,
        etag=state.get("etag", ""),
        last_run_date=state.get("last_run_date", ""),
    )
    print(f"  Migrated: last_run_date={state.get('last_run_date')}")


def migrate_version_cache(dept: str, comparison_xlsx: Path):
    print("\n[3] Migrating version cache...")
    if not comparison_xlsx.exists():
        print("  No latest_comparison.xlsx found — skipping")
        return
    try:
        wb      = load_workbook(str(comparison_xlsx), read_only=True, data_only=True)
        ws      = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows    = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            r   = dict(zip(headers, row))
            nct = str(r.get("NCT ID", "") or "").strip()
            if nct:
                rows.append({
                    "nct_id":             nct,
                    "note":               str(r.get("Note", "") or ""),
                    "total_versions":     r.get("Total Versions"),
                    "prev_version":       r.get("Prev Version"),
                    "curr_version":       r.get("Curr Version"),
                    "prev_date":          str(r.get("Prev Date", "") or ""),
                    "curr_date":          str(r.get("Curr Date", "") or ""),
                    "curr_status":        str(r.get("Current Status", "") or ""),
                    "modules_changed":    str(r.get("Modules Changed", "") or ""),
                    "field_change_count": r.get("Field Change Count"),
                    "field_changes":      str(r.get("Field Changes", "") or ""),
                })
        wb.close()
        db.upsert_version_cache(dept, rows)
        print(f"  Migrated {len(rows)} version cache rows")
    except Exception as e:
        print(f"  ERROR: {e}")


def migrate_organized_data(dept: str, organized_xlsx: Path):
    print("\n[4] Migrating organized trial data...")
    if not organized_xlsx.exists():
        print("  No Clinical_Trials_Organized.xlsx found — skipping")
        return
    try:
        df   = pd.read_excel(organized_xlsx, engine="openpyxl", dtype=str)
        rows = df.fillna("").to_dict(orient="records")
        db.upsert_organized_trials(dept, rows)
        print(f"  Migrated {len(rows)} organized trial rows")
    except Exception as e:
        print(f"  ERROR: {e}")


def migrate_combined_output(dept: str, output_dir: Path):
    print("\n[5] Migrating historical combined output files...")
    files = sorted(output_dir.glob("combined_*.xlsx"))
    if not files:
        print("  No combined output files found — skipping")
        return

    for xlsx in files:
        run_date = xlsx.stem.replace("combined_", "")[:10]  # YYYY-MM-DD
        print(f"  Processing {xlsx.name} (run_date={run_date})...")
        try:
            wb = load_workbook(str(xlsx), read_only=True, data_only=True)

            def _read(sheet_name: str) -> list[dict]:
                aliases = {
                    "New ADC Trials Found":        ["New ADC Trials Found", "New Candidates"],
                    "Unmatched Trials - Gap Review": ["Unmatched Trials - Gap Review", "Unmatched Trials"],
                    "CT.gov Updated This Week":    ["CT.gov Updated This Week", "Modified Trials"],
                    "Field-Level Changes":         ["Field-Level Changes", "New Updates This Run", "Field Changes"],
                }
                candidates = aliases.get(sheet_name, [sheet_name])
                ws = next((wb[n] for n in candidates if n in wb.sheetnames), None)
                if not ws:
                    return []
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1), [])]
                return [
                    dict(zip(headers, row))
                    for row in ws.iter_rows(min_row=2, values_only=True)
                    if any(v for v in row)
                ]

            new_candidates  = _read("New ADC Trials Found")
            unmatched       = _read("Unmatched Trials - Gap Review")
            modified        = _read("CT.gov Updated This Week")
            field_changes_raw = _read("Field-Level Changes")

            # Map field changes to pipeline row format
            field_changes = [{
                "nct_id":             str(r.get("NCT ID", "") or ""),
                "note":               str(r.get("Note", "") or ""),
                "total_versions":     r.get("Total Versions"),
                "prev_version":       r.get("Prev Version"),
                "curr_version":       r.get("Curr Version"),
                "prev_date":          str(r.get("Prev Date", "") or ""),
                "curr_date":          str(r.get("Curr Date", "") or ""),
                "curr_status":        str(r.get("Current Status", "") or ""),
                "modules_changed":    str(r.get("Modules Changed", "") or ""),
                "field_change_count": r.get("Field Change Count"),
                "field_changes":      str(r.get("Field Changes", "") or ""),
            } for r in field_changes_raw if r.get("NCT ID")]

            wb.close()

            if new_candidates:
                db.insert_new_candidates(dept, run_date, new_candidates)
            if unmatched:
                db.upsert_unmatched(dept, run_date, unmatched)
            if modified:
                db.insert_modified_log(dept, run_date, modified)
            if field_changes:
                db.insert_field_changes(dept, run_date, field_changes)

            # Run history row
            db.insert_run_history(
                dept=dept, run_date=run_date,
                new_candidates=len(new_candidates),
                unmatched=len(unmatched),
                modified_trials=len(modified),
                field_diffs=len(field_changes),
                canonical_fixes=0,
                duration_s=0,
                output_file=xlsx.name,
            )
            print(f"    New={len(new_candidates)} Unmatched={len(unmatched)} "
                  f"Modified={len(modified)} FieldChanges={len(field_changes)}")

        except Exception as e:
            print(f"  ERROR processing {xlsx.name}: {e}")


def main():
    parser = argparse.ArgumentParser(description="Migrate Excel data to Supabase")
    parser.add_argument("dept", help='Department name, e.g. "ADC"')
    args = parser.parse_args()

    dept      = args.dept
    _here     = Path(__file__).parent
    dept_path = _here / dept
    output_dir = dept_path / "output"

    print(f"\n{'='*55}")
    print(f"NCT Tracking — Migration to Supabase")
    print(f"Department : {dept}")
    print(f"{'='*55}")

    # Auto-detect tracking list
    skip = {"latest_comparison.xlsx", f"{dept}_Master_Dashboard.xlsx"}
    xlsx_files = [
        f for f in dept_path.glob("*.xlsx")
        if not f.name.startswith("~") and f.name not in skip
        and not f.name.endswith("_Master_Dashboard.xlsx")
    ]
    if not xlsx_files:
        print("ERROR: No tracking list .xlsx found in dept folder.")
        sys.exit(1)
    tracking_xlsx = xlsx_files[0]
    print(f"Tracking list : {tracking_xlsx.name}")

    migrate_tracking_list(dept, tracking_xlsx)
    migrate_pipeline_state(dept, dept_path / "state.json")
    migrate_version_cache(dept, dept_path / "latest_comparison.xlsx")
    migrate_organized_data(dept, output_dir / "Clinical_Trials_Organized.xlsx")
    migrate_combined_output(dept, output_dir)

    print(f"\n{'='*55}")
    print("Migration complete. Verify in Supabase Dashboard -> Table Editor.")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
