"""
Master Dashboard Updater

Reads the latest combined_*.xlsx output and Clinical_Trials_Organized.xlsx,
then appends/upserts into the persistent <DEPT>_Master_Dashboard.xlsx.

Usage (standalone):
    python master_dashboard.py ADC

Called automatically from combined_pipeline.py at the end of every run.
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Styling (mirrors combined_pipeline.py) ────────────────────────────────────
_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

FILL_CANDIDATE = PatternFill("solid", fgColor="C0392B")   # red
FILL_MODIFIED  = PatternFill("solid", fgColor="D35400")   # orange
FILL_UNMATCHED = PatternFill("solid", fgColor="7B68EE")   # purple
FILL_CHANGES   = PatternFill("solid", fgColor="117A65")   # teal
FILL_CANONICAL = PatternFill("solid", fgColor="555555")   # gray
FILL_ORGANIZED = PatternFill("solid", fgColor="1A5276")   # dark blue
HDR_FONT       = Font(color="FFFFFF", bold=True)

# ── Column definitions ────────────────────────────────────────────────────────
_DISCOVERY_COLS = [
    "Run Date", "Decision",
    "NCT ID", "Title", "Conditions", "Interventions",
    "Sponsor", "Recruitment Status", "Study First Posted", "Last Updated", "Link",
]
_CHANGES_COLS = [
    "Run Date", "NCT ID", "Note", "Total Versions", "Prev Version", "Curr Version",
    "Prev Date", "Curr Date", "Current Status", "Modules Changed",
    "Field Change Count", "Field Changes",
]
_CANONICAL_COLS = ["Run Date", "Requested NCT ID", "Canonical NCT ID", "Already Tracked"]
_RUNHIST_COLS   = [
    "Run Date", "New Candidates", "Unmatched", "Modified Trials",
    "Field Diffs", "Canonical Fixes", "Duration (s)", "Output File",
]

_COL_WIDTHS = {
    "Run Date": 14, "Decision": 14, "NCT ID": 14,
    "Requested NCT ID": 14, "Canonical NCT ID": 14,
    "Title": 55, "Conditions": 35, "Interventions": 35,
    "Field Changes": 60, "Sponsor": 30,
    "Recruitment Status": 22, "Study First Posted": 16,
    "Last Updated": 14, "Link": 45, "Output File": 45,
}

# Master workbook sheet names
_SHEET_NEW_TRIALS    = "New ADC Trials Found"
_SHEET_UNMATCHED     = "Unmatched Trials"
_SHEET_FIELD_CHANGES = "Field Changes History"
_SHEET_CANONICAL     = "Canonical History"
_SHEET_RUN_HIST      = "Run History"
_SHEET_ORGANIZED     = "All Tracked Trials (Current)"

# Source sheet names in combined output Excel (must match combined_pipeline SHEET_* constants)
_SRC_NEW_TRIALS    = "New ADC Trials Found"
_SRC_UNMATCHED     = "Unmatched Trials - Gap Review"
_SRC_MODIFIED      = "CT.gov Updated This Week"
_SRC_FIELD_CHANGES = "Field-Level Changes"
_SRC_REDIRECTS     = "Redirected NCT IDs"


# ── Sheet helpers ─────────────────────────────────────────────────────────────

def _init_sheet(ws, cols: list[str], fill: PatternFill):
    """Write styled header row if the sheet is empty."""
    if ws.cell(1, 1).value is not None:
        return  # already initialised
    for col_idx, col in enumerate(cols, 1):
        cell            = ws.cell(row=1, column=col_idx, value=col)
        cell.fill       = fill
        cell.font       = HDR_FONT
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col, 20)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    if "Decision" in cols:
        d_letter = get_column_letter(cols.index("Decision") + 1)
        dv = DataValidation(
            type="list", formula1='"Pending,Approved,Rejected,Tracked"',
            allow_blank=True, showDropDown=False,
        )
        dv.sqref = f"{d_letter}2:{d_letter}100000"
        ws.add_data_validation(dv)


def _append_rows(ws, rows: list[dict], cols: list[str]):
    """Append rows below existing data."""
    if not rows:
        return
    next_row = ws.max_row + 1
    for row_dict in rows:
        for col_idx, col in enumerate(cols, 1):
            val  = row_dict.get(col, "")
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border    = _BORDER
            if col == "Link" and str(val or "").startswith("http"):
                cell.hyperlink = str(val)
                cell.font = Font(color="0563C1", underline="single")
        next_row += 1


def _update_tracked_decisions(ws, tracked_set: set[str]):
    """Flip Decision Pending → Tracked for NCTs now in the tracking list."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        nct_col  = headers.index("NCT ID")   + 1
        dec_col  = headers.index("Decision") + 1
    except ValueError:
        return
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, dec_col).value or "").strip().lower() == "pending":
            nct = str(ws.cell(row, nct_col).value or "").strip()
            if nct in tracked_set:
                ws.cell(row, dec_col).value = "Tracked"


def _replace_sheet(wb, name: str, fill: PatternFill, rows: list[dict]):
    """Delete and recreate a sheet with 60+ col organized data."""
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    if not rows:
        return

    all_keys = next((list(r.keys()) for r in rows if r and len(r) > 1), ["NCT ID"])
    for col_idx, col in enumerate(all_keys, 1):
        cell            = ws.cell(row=1, column=col_idx, value=col)
        cell.fill       = fill
        cell.font       = HDR_FONT
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col, 22)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_keys))}1"

    for row_idx, row_dict in enumerate(rows, 2):
        for col_idx, col in enumerate(all_keys, 1):
            val  = row_dict.get(col, "") if row_dict else ""
            val  = "" if val is None else val
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border    = _BORDER


# ── Read from combined output Excel ──────────────────────────────────────────

def _read_sheet_as_dicts(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []
    ws      = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1), [])]
    rows    = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if any(v for v in row_dict.values()):
            rows.append(row_dict)
    return rows


# ── Main update function ──────────────────────────────────────────────────────

def update_master(
    dept_dir:    Path,
    output_dir:  Path,
    run_date:    str      = None,
    duration_s:  float    = None,
    latest_file: Path     = None,
):
    """
    Update <dept>_Master_Dashboard.xlsx from the latest combined output.

    dept_dir    : department folder (e.g. ADC/)
    output_dir  : dept/output/ folder
    run_date    : YYYY-MM-DD string; defaults to today
    duration_s  : run duration in seconds (None when called standalone)
    latest_file : path to combined_*.xlsx; auto-detected if None
    """
    run_date = run_date or datetime.today().strftime("%Y-%m-%d")

    # Auto-detect latest combined output
    if latest_file is None:
        files = sorted(output_dir.glob("combined_*.xlsx"), reverse=True)
        if not files:
            print("  No combined output files found — nothing to update.")
            return
        latest_file = files[0]

    print(f"  Reading: {latest_file.name}")

    wb_out = load_workbook(str(latest_file), read_only=True, data_only=True)

    new_candidates  = _read_sheet_as_dicts(wb_out, _SRC_NEW_TRIALS)
    unmatched       = _read_sheet_as_dicts(wb_out, _SRC_UNMATCHED)
    modified_trials = _read_sheet_as_dicts(wb_out, _SRC_MODIFIED)
    redirects_raw   = _read_sheet_as_dicts(wb_out, _SRC_REDIRECTS)
    field_changes   = _read_sheet_as_dicts(wb_out, _SRC_FIELD_CHANGES)
    wb_out.close()

    # Load organized data (All Tracked Trials sheet)
    organized_rows: list[dict] = []
    org_path = output_dir / "Clinical_Trials_Organized.xlsx"
    if org_path.exists():
        wb_org = load_workbook(str(org_path), read_only=True, data_only=True)
        organized_rows = _read_sheet_as_dicts(wb_org, wb_org.sheetnames[0])
        wb_org.close()

    # Determine current tracked NCT IDs from organized data
    tracked_set = {str(r.get("NCT ID", "")) for r in organized_rows if r.get("NCT ID")}

    # Open or create master workbook
    master_path = dept_dir / f"{dept_dir.name}_Master_Dashboard.xlsx"
    if master_path.exists():
        wb = load_workbook(str(master_path))
    else:
        wb = Workbook()
        wb.active.title = _SHEET_NEW_TRIALS

    def _get(name):
        return wb[name] if name in wb.sheetnames else wb.create_sheet(name)

    # ── Sheet 1: New ADC Trials Found ────────────────────────────────────────
    ws1 = _get(_SHEET_NEW_TRIALS)
    _init_sheet(ws1, _DISCOVERY_COLS, FILL_CANDIDATE)
    _update_tracked_decisions(ws1, tracked_set)
    _append_rows(ws1, [
        {"Run Date": run_date,
         "Decision": "Tracked" if r.get("NCT ID", "") in tracked_set else "Pending",
         **{c: r.get(c, "") for c in _DISCOVERY_COLS[2:]}}
        for r in new_candidates
    ], _DISCOVERY_COLS)

    # ── Sheet 2: Unmatched Trials ─────────────────────────────────────────────
    ws2 = _get(_SHEET_UNMATCHED)
    _init_sheet(ws2, _DISCOVERY_COLS, FILL_UNMATCHED)
    _update_tracked_decisions(ws2, tracked_set)
    _append_rows(ws2, [
        {"Run Date": run_date, "Decision": "Pending",
         **{c: r.get(c, "") for c in _DISCOVERY_COLS[2:]}}
        for r in unmatched
    ], _DISCOVERY_COLS)

    # ── Sheet 3: Field Changes History ────────────────────────────────────────
    ws3 = _get(_SHEET_FIELD_CHANGES)
    _init_sheet(ws3, _CHANGES_COLS, FILL_CHANGES)
    _append_rows(ws3, [
        {
            "Run Date":           run_date,
            "NCT ID":             r.get("NCT ID", ""),
            "Note":               r.get("Note", ""),
            "Total Versions":     r.get("Total Versions", ""),
            "Prev Version":       r.get("Prev Version", ""),
            "Curr Version":       r.get("Curr Version", ""),
            "Prev Date":          r.get("Prev Date", ""),
            "Curr Date":          r.get("Curr Date", ""),
            "Current Status":     r.get("Current Status", ""),
            "Modules Changed":    r.get("Modules Changed", ""),
            "Field Change Count": r.get("Field Change Count", ""),
            "Field Changes":      r.get("Field Changes", ""),
        }
        for r in field_changes
    ], _CHANGES_COLS)

    # ── Sheet 4: Canonical History ────────────────────────────────────────────
    ws4 = _get(_SHEET_CANONICAL)
    _init_sheet(ws4, _CANONICAL_COLS, FILL_CANONICAL)
    _append_rows(ws4, [
        {
            "Run Date":         run_date,
            "Requested NCT ID": r.get("Requested NCT ID", ""),
            "Canonical NCT ID": r.get("Canonical NCT ID", ""),
            "Already Tracked":  str(r.get("Already Tracked", "")),
        }
        for r in redirects_raw
    ], _CANONICAL_COLS)

    # ── Sheet 5: Run History ──────────────────────────────────────────────────
    ws5 = _get(_SHEET_RUN_HIST)
    _init_sheet(ws5, _RUNHIST_COLS, FILL_MODIFIED)
    _append_rows(ws5, [{
        "Run Date":        run_date,
        "New Candidates":  len(new_candidates),
        "Unmatched":       len(unmatched),
        "Modified Trials": len(modified_trials),
        "Field Diffs":     len(field_changes),
        "Canonical Fixes": len(redirects_raw),
        "Duration (s)":    round(duration_s, 1) if duration_s is not None else "",
        "Output File":     latest_file.name,
    }], _RUNHIST_COLS)

    # ── Sheet 6: All Tracked Trials (Current) — replace ──────────────────────
    _replace_sheet(wb, _SHEET_ORGANIZED, FILL_ORGANIZED, organized_rows)

    wb.save(str(master_path))
    print(f"  Master Dashboard saved: {master_path.name}")
    print(f"    New ADC Trials  : +{len(new_candidates)}")
    print(f"    Unmatched       : +{len(unmatched)}")
    print(f"    Field Changes   : +{len(field_changes)}")
    print(f"    Run History     : +1 row  ({run_date})")
    print(f"    All Tracked     : {len(organized_rows)} rows (current)")


# ── CLI entry point ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Update ADC Master Dashboard")
    parser.add_argument("dept", help='Department folder name (e.g. "ADC") or full path')
    args = parser.parse_args()

    _here     = Path(__file__).parent
    dept_path = Path(args.dept) if Path(args.dept).is_absolute() else _here / args.dept
    if not dept_path.exists():
        print(f"ERROR: Department folder not found: {dept_path}")
        sys.exit(1)

    output_dir = dept_path / "output"
    update_master(dept_dir=dept_path, output_dir=output_dir)


if __name__ == "__main__":
    main()
