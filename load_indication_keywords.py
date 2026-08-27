"""
Load indication keywords from Excel files into dept_indications table.
Keywords are comma-separated in Excel; stored as pipe-separated (|) in DB.
"""
import openpyxl
import psycopg2
import psycopg2.extras
import os
from pathlib import Path
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")
DB_URL = os.environ.get("DATABASE_URL", "")

ADC_FILE  = r"C:\Users\LAPTOP\Downloads\Keyword list_SAM_BeOne_WHTx.xlsx"
ASMB_FILE = r"C:\Users\LAPTOP\Downloads\ASMB Overall List_DN.xlsx"


def comma_to_pipe(raw: str) -> str:
    parts = [k.strip() for k in raw.split(",") if k.strip()]
    return "|".join(parts)


def read_adc():
    wb = openpyxl.load_workbook(ADC_FILE)
    ws = wb["Indications"]
    rows = []
    for indication, keywords in ws.iter_rows(min_row=2, values_only=True):
        if indication and keywords:
            rows.append(("ADC", str(indication).strip(), comma_to_pipe(str(keywords))))
    return rows


def read_asmb():
    wb = openpyxl.load_workbook(ASMB_FILE)
    ws = wb["All Indications"]
    rows = []
    for indication, synonyms in ws.iter_rows(min_row=2, values_only=True):
        if indication and synonyms:
            rows.append(("ASMB", str(indication).strip(), comma_to_pipe(str(synonyms))))
    return rows


def upsert(conn, records):
    with conn.cursor() as cur:
        cur.execute('SET search_path TO "CT"')
        for dept, indication, keywords in records:
            cur.execute(
                """
                INSERT INTO dept_indications (dept, indication, keywords)
                VALUES (%s, %s, %s)
                ON CONFLICT (dept, indication)
                DO UPDATE SET keywords = EXCLUDED.keywords
                """,
                (dept, indication, keywords),
            )
    conn.commit()


def main():
    if not DB_URL:
        raise RuntimeError("DATABASE_URL env var not set")

    adc_rows  = read_adc()
    asmb_rows = read_asmb()

    print(f"ADC indications: {len(adc_rows)}")
    for dept, ind, kw in adc_rows:
        print(f"  [{dept}] {ind}: {kw[:80]}...")

    print(f"\nASMB indications: {len(asmb_rows)}")
    for dept, ind, kw in asmb_rows:
        print(f"  [{dept}] {ind}: {kw[:80]}...")

    conn = psycopg2.connect(DB_URL)
    try:
        upsert(conn, adc_rows + asmb_rows)
        print(f"\nInserted/updated {len(adc_rows) + len(asmb_rows)} rows into dept_indications.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
