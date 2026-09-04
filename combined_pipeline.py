"""
NCT Combined Tracking Pipeline — Multi-Department

Usage:
    python combined_pipeline.py ADC
    python combined_pipeline.py ASMB
    python combined_pipeline.py "C:\\full\\path\\to\\dept"

Department folder structure:
    <dept>\
        keywords\
            general_terms.txt             one keyword per line, # = comment
            ADC_Assets_Alias.xlsx         Drug Name / Alias Name columns
            any_other.xlsx or .txt        any additional keyword source
        <TrackingList>.xlsx               only one .xlsx in the dept root
        output\                           auto-created
        state.json                        auto-created
        latest_comparison.xlsx            auto-created (Flow 2 version cache)

Keyword matching:
    Drug Name + pipe-split Alias Names are each treated as individual keywords.
    Alias "-" is ignored. Matching is exact case-insensitive substring.
    "9Bio Therapeutics ADC" matches as one phrase, not split into tokens.

Flows:
    FLOW 1 — Global Discovery
        REST API scan (LastUpdatePostDate range) -> ~8,575 NCTs per Thursday
        In tracking list + date changed  -> MODIFIED
        In tracking list + date same     -> STALE (skip)
        Not in list + keyword match      -> NEW CANDIDATE
        Not in list + no match           -> UNMATCHED (sent for keyword gap review)

    FLOW 2 — Field-Level Diff
        Internal API version check per tracked NCT
        Same version  -> Skip (carry forward)
        New version   -> Full field-level diff
        Different ID  -> Canonical auto-replace in tracking list

Output Excel (6 sheets):
    1. New Candidates        keyword matched, pending team approval
    2. Modified Trials       already tracked, date changed
    3. Unmatched Trials      no keyword match, for gap review
    4. Field Changes         what changed per tracked NCT (full snapshot)
    5. New Updates This Run  subset of field changes — only new versions
    6. Canonical NCTs        redirected IDs, auto-fixed
"""

import argparse
import json
import pandas as pd
import re
import smtplib
import sys
import time
import requests
from datetime import datetime, timedelta, date
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Import from existing NCT scripts ──────────────────────────────────────────
_NCT_DIR = Path(r"C:\Users\LAPTOP\Downloads\NCT_Record_History_Changes")
sys.path.insert(0, str(_NCT_DIR))

import db
import norm

from NCT_Changes_Tracker import (
    get_history_list,
    get_latest_comparison,
    get_version_data,
    compare_versions,
    format_field_changes,
)
from all_versions_compare import get_all_comparisons as _get_all_version_pairs
from CT_PULL import fetch_study_data, build_raw_row

# ── Organized data parsing (62-column schema from CT.gov raw API JSON) ─────────

def _nest(obj: dict, *keys, default=""):
    curr = obj
    for k in keys:
        if isinstance(curr, dict) and k in curr:
            curr = curr[k]
        else:
            return default
    return curr if curr is not None else default


def _fmt_outcomes(lst: list) -> str:
    return "\n".join(
        f"{i}. {o.get('measure','N/A')} [Timeframe: {o.get('timeFrame','No timeframe')}]"
        for i, o in enumerate(lst, 1)
    ) if lst else ""


def _fmt_contacts(lst: list) -> str:
    parts = []
    for c in lst:
        s = f"{c.get('name','')} ({c.get('role','')})"
        if c.get("phone") or c.get("email"):
            s += f" - Phone: {c.get('phone','')}, Email: {c.get('email','')}"
        parts.append(s)
    return "\n".join(parts)


def _fmt_officials(lst: list) -> str:
    return "\n".join(
        f"{o.get('name','')} ({o.get('role','')}) - {o.get('affiliation','')}"
        for o in lst
    )


def _fmt_locations(locs: list) -> str:
    if not locs:
        return ""
    grouped: dict = {}
    for loc in locs:
        country = loc.get("country", "Unknown")
        state   = loc.get("state", "Other")
        grouped.setdefault(country, {}).setdefault(state, [])
        parts = [loc.get("city"), loc.get("state"), loc.get("country"), loc.get("zip")]
        addr  = ", ".join(str(p) for p in parts if p)
        grouped[country][state].append((addr, loc.get("facility", "")))
    lines = [f"Total Locations: {len(locs)}", "", "--- Locations ---"]
    for i, country in enumerate(sorted(grouped), 1):
        lines.append(f"{i}) {country}:")
        for state in sorted(grouped[country]):
            lines.append(f"   - {state}:")
            for addr, fac in grouped[country][state]:
                lines.append(f"       {addr}")
                if fac:
                    lines.append(f"       {fac}")
            lines.append("")
    return "\n".join(lines).strip()


def _extract_countries(locs: list) -> str:
    """Return pipe-separated unique country names from raw CT.gov location list."""
    seen: list[str] = []
    for loc in locs:
        c = (loc.get("country") or "").strip()
        if c and c not in seen:
            seen.append(c)
    return " | ".join(sorted(seen))


def _extract_sites(locs: list) -> str:
    """Return pipe-separated unique facility names, stripping /ID# suffixes."""
    seen: list[str] = []
    for loc in locs:
        fac = (loc.get("facility") or "").strip()
        fac = re.sub(r'\s*/ID#\s*\S+', '', fac).strip()
        fac = re.sub(r'^\d+\.\s+', '', fac).strip()  # strip leading "1. " numbering
        # Skip pure numeric site codes (e.g. Chinese trials storing "015", "016")
        if not fac or not re.search(r'[A-Za-z]', fac):
            continue
        if fac not in seen:
            seen.append(fac)
    return " | ".join(seen)


ORGANIZED_COLS = [
    "NCT ID", "Study URL", "Other IDs",
    "Brief Title", "Official Title", "Acronym", "Org Full Name",
    "Overall Status", "Status Verified Date", "Exclusion Rationale",
    "Expanded Access Info",
    "Start Date", "Start DateType",
    "Primary Completion Date", "Primary Completion DateType",
    "completionDateStructDate", "completionDateStructType",
    "studyFirstSubmitDate", "studyFirstSubmitQcDate",
    "Study First Post Date", "studyFirstPostDateType",
    "Last Update Submit Date", "Last Update Post Date", "lastUpdatePostDateType",
    "Sponsors", "Collaborators", "Funder Type",
    "responsiblePartyType", "responsiblePartyleadSponsor",
    "responsiblePartyleadSponsorclass",
    "FDA Regulated Drug", "FDA Regulated Device", "Has DMC",
    "briefSummary", "detailedDescription",
    "conditions", "studyType", "phases",
    "Allocation", "Intervention Model", "Masking", "Primary Purpose",
    "Enrollment", "enrollmentInfoType", "Interventions",
    "Primary Outcomes", "Secondary Outcomes",
    "eligibilityCriteria", "Sex", "Minimum Age", "Standard Ages",
    "healthyVolunteers",
    "Central Contacts", "Overall Officials", "Locations",
    "countries", "sites",
    "Study Documents", "Reference Count", "PMIDs", "Citations",
    "IPD Sharing", "MeSH Conditions", "MeSH Interventions",
    "Primary Drug",
    "_parsed_date",
]


def parse_study_to_organized(nct_id: str, data: dict) -> dict:
    """Parse raw CT.gov API JSON into the full 62-column organized schema."""
    p  = data.get("protocolSection", {})
    d  = data.get("derivedSection", {})

    ident  = p.get("identificationModule", {})
    st     = p.get("statusModule", {})
    sp     = p.get("sponsorCollaboratorsModule", {})
    ov     = p.get("oversightModule", {})
    desc   = p.get("descriptionModule", {})
    cond_m = p.get("conditionsModule", {})
    design = p.get("designModule", {})
    arms   = p.get("armsInterventionsModule", {})
    outc   = p.get("outcomesModule", {})
    elig   = p.get("eligibilityModule", {})
    cl     = p.get("contactsLocationsModule", {})
    refs_m = p.get("referencesModule", {})
    ipd    = p.get("ipdSharingStatementModule", {})
    cb     = d.get("conditionBrowseModule", {})
    ib     = d.get("interventionBrowseModule", {})

    nct       = _nest(ident, "nctId") or nct_id
    org_id    = _nest(ident, "orgStudyIdInfo", "id")
    sec_ids   = [str(s.get("id", "")) for s in ident.get("secondaryIdInfos", []) if s.get("id")]
    other_ids = ", ".join(([str(org_id)] if org_id else []) + sec_ids)

    interventions_list = arms.get("interventions", [])
    interventions_str  = ", ".join(i.get("name", "") for i in interventions_list)

    refs      = refs_m.get("references", [])
    pmids     = [str(r.get("pmid", "")) for r in refs if r.get("pmid")]
    citations = [f"[{r.get('type','UNKNOWN')}] {r.get('citation','')}" for r in refs if r.get("citation")]

    docs = []
    for doc in d.get("largeDocModule", {}).get("largeDocs", []):
        fn = doc.get("filename", "")
        if fn and nct:
            url = f"https://cdn.clinicaltrials.gov/large-docs/{nct[-2:]}/{nct}/{fn}"
            docs.append(f"{doc.get('label','')}, {url}")

    return {
        "NCT ID":         nct,
        "Study URL":      f"https://clinicaltrials.gov/study/{nct}" if nct else "",
        "Other IDs":      other_ids,
        "Brief Title":    _nest(ident, "briefTitle"),
        "Official Title": _nest(ident, "officialTitle"),
        "Acronym":        _nest(ident, "acronym"),
        "Org Full Name":  _nest(ident, "organization", "fullName"),
        "Overall Status":              _nest(st, "overallStatus"),
        "Status Verified Date":        _nest(st, "statusVerifiedDate"),
        "Exclusion Rationale":         _nest(st, "whyStopped"),
        "Expanded Access Info":        str(_nest(st, "expandedAccessInfo", "hasExpandedAccess")),
        "Start Date":                  _nest(st, "startDateStruct", "date"),
        "Start DateType":              _nest(st, "startDateStruct", "type"),
        "Primary Completion Date":     _nest(st, "primaryCompletionDateStruct", "date"),
        "Primary Completion DateType": _nest(st, "primaryCompletionDateStruct", "type"),
        "completionDateStructDate":    _nest(st, "completionDateStruct", "date"),
        "completionDateStructType":    _nest(st, "completionDateStruct", "type"),
        "studyFirstSubmitDate":        _nest(st, "studyFirstSubmitDate"),
        "studyFirstSubmitQcDate":      _nest(st, "studyFirstSubmitQcDate"),
        "Study First Post Date":       _nest(st, "studyFirstPostDateStruct", "date"),
        "studyFirstPostDateType":      _nest(st, "studyFirstPostDateStruct", "type"),
        "Last Update Submit Date":     _nest(st, "lastUpdateSubmitDate"),
        "Last Update Post Date":       _nest(st, "lastUpdatePostDateStruct", "date"),
        "lastUpdatePostDateType":      _nest(st, "lastUpdatePostDateStruct", "type"),
        "Sponsors":                         norm.normalize_org(_nest(sp, "leadSponsor", "name") or ""),
        "Collaborators":                    norm.normalize_orgs_field(" | ".join(c.get("name", "") for c in sp.get("collaborators", []))),
        "Funder Type":                      _nest(sp, "leadSponsor", "class"),
        "responsiblePartyType":             _nest(sp, "responsibleParty", "type"),
        "responsiblePartyleadSponsor":      norm.normalize_org(_nest(sp, "leadSponsor", "name") or ""),
        "responsiblePartyleadSponsorclass": _nest(sp, "leadSponsor", "class"),
        "FDA Regulated Drug":   str(_nest(ov, "isFdaRegulatedDrug")),
        "FDA Regulated Device": str(_nest(ov, "isFdaRegulatedDevice")),
        "Has DMC":              str(_nest(ov, "hasDmc")),
        "briefSummary":        _nest(desc, "briefSummary"),
        "detailedDescription": _nest(desc, "detailedDescription"),
        "conditions":         norm.normalize_conditions_field(", ".join(cond_m.get("conditions", []))),
        # Primary Drug is tagged separately via dept_keywords (already canonical)
        "studyType":          _nest(design, "studyType"),
        "phases":             ", ".join(design.get("phases", [])),
        "Allocation":         _nest(design, "designInfo", "allocation"),
        "Intervention Model": _nest(design, "designInfo", "interventionModel"),
        "Masking":            _nest(design, "designInfo", "maskingInfo", "masking"),
        "Primary Purpose":    _nest(design, "designInfo", "primaryPurpose"),
        "Enrollment":         str(_nest(design, "enrollmentInfo", "count")),
        "enrollmentInfoType": _nest(design, "enrollmentInfo", "type"),
        "Interventions":      interventions_str,
        "Primary Outcomes":   _fmt_outcomes(outc.get("primaryOutcomes", [])),
        "Secondary Outcomes": _fmt_outcomes(outc.get("secondaryOutcomes", [])),
        "eligibilityCriteria": _nest(elig, "eligibilityCriteria"),
        "Sex":                 _nest(elig, "sex"),
        "Minimum Age":         _nest(elig, "minimumAge"),
        "Standard Ages":       ", ".join(elig.get("stdAges", [])),
        "healthyVolunteers":   str(_nest(elig, "healthyVolunteers")),
        "Central Contacts":    _fmt_contacts(cl.get("centralContacts", [])),
        "Overall Officials":   _fmt_officials(cl.get("overallOfficials", [])),
        "Locations":           _fmt_locations(cl.get("locations", [])),
        "countries":           _extract_countries(cl.get("locations", [])),
        "sites":               _extract_sites(cl.get("locations", [])),
        "IPD Sharing":         _nest(ipd, "ipdSharing"),
        "MeSH Conditions":     " | ".join(
            f"{m.get('id','')}: {m.get('term','')}"
            for m in cb.get("meshes", []) if m.get("term")
        ),
        "MeSH Interventions":  " | ".join(
            f"{m.get('id','')}: {m.get('term','')}"
            for m in ib.get("meshes", []) if m.get("term")
        ),
        "Study Documents":  " | ".join(docs),
        "Reference Count":  len(refs),
        "PMIDs":            ", ".join(pmids),
        "Citations":        " | ".join(citations),
        "Primary Drug":     "",
        "_parsed_date":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ── Primary Drug keyword matching ──────────────────────────────────────────────

def load_drug_keywords() -> list[tuple[str, list[str]]]:
    """Load drug names and aliases from dept_keywords table."""
    try:
        drugs: list[tuple[str, list[str]]] = []
        for row in db.load_dept_keywords(DEPT_NAME):
            name = str(row.get("drug_name", "") or "").strip()
            if not name or name in ("nan", "-"):
                continue
            alias_raw = str(row.get("alias_names", "") or "").strip()
            aliases: list[str] = []
            if alias_raw and alias_raw not in ("-", "nan", ""):
                for a in alias_raw.split("|"):
                    a = a.strip()
                    if a and a not in ("-", ""):
                        aliases.append(a)
            drugs.append((name, aliases))
        print(f"  [Drug Tag] Loaded {len(drugs)} drug keywords from DB")
        return drugs
    except Exception as e:
        print(f"  [Drug Tag] Warning: could not load drug keywords from DB: {e}")
        return []


def _compile_drug_patterns(drugs: list[tuple[str, list[str]]]) -> list[tuple[str, re.Pattern]]:
    """Pre-compile one regex per drug (all its aliases ORed together). Call once, reuse many times."""
    compiled = []
    for drug_name, aliases in drugs:
        terms = [t for t in ([drug_name] + aliases) if t and t != "-"]
        if not terms:
            continue
        pattern = "|".join(r"(?<![A-Za-z0-9])" + re.escape(t) + r"(?![A-Za-z0-9])" for t in terms)
        compiled.append((drug_name, re.compile(pattern, re.IGNORECASE)))
    return compiled


def tag_primary_drug(interventions: str, drugs: list[tuple[str, list[str]]] | list[tuple[str, re.Pattern]]) -> str:
    """
    Exact word-boundary match of drug names/aliases against the Interventions text.
    Accepts either raw drug list or pre-compiled patterns from _compile_drug_patterns().
    Multiple matched Drug Names are joined with ' | '.
    """
    if not interventions or not drugs:
        return ""
    matched: list[str] = []
    for drug_name, term_or_pattern in drugs:
        if isinstance(term_or_pattern, re.Pattern):
            if term_or_pattern.search(interventions):
                matched.append(drug_name)
        else:
            for term in ([drug_name] + term_or_pattern):
                if not term or term == "-":
                    continue
                pattern = r"(?<![A-Za-z0-9])" + re.escape(term) + r"(?![A-Za-z0-9])"
                if re.search(pattern, interventions, re.IGNORECASE):
                    matched.append(drug_name)
                    break
    return " | ".join(dict.fromkeys(matched))  # deduplicate, preserve match order


def tag_all_organized_drug(dept: str, drugs: list[tuple[str, list[str]]]) -> None:
    """
    Re-tag Primary Drug for every organized trial of dept.
    Reads Interventions from Supabase → computes matches → batch-updates Primary Drug.
    Safe to re-run whenever the keywords file changes.
    """
    if not drugs:
        print("  [Drug Tag] No drug keywords loaded — skipping")
        return
    interventions_map = db.load_all_interventions(dept)
    if not interventions_map:
        print("  [Drug Tag] No organized trials found")
        return
    compiled = _compile_drug_patterns(drugs)
    updates = {
        nct_id: tag_primary_drug(iv, compiled)
        for nct_id, iv in interventions_map.items()
    }
    db.batch_update_primary_drug(dept, updates)
    tagged = sum(1 for v in updates.values() if v)
    print(f"  [Drug Tag] Tagged {tagged}/{len(updates)} trials with Primary Drug")


# ── API constants (shared across departments) ──────────────────────────────────
RSS_URL     = "https://clinicaltrials.gov/api/rss?dateField=LastUpdatePostDate"
API_URL     = "https://clinicaltrials.gov/api/v2/studies"
PUB_HEADERS = {"User-Agent": "TrialsTracker/1.0"}

SCAN_FIELDS = ",".join([
    # Identification
    "protocolSection.identificationModule.nctId",
    "protocolSection.identificationModule.briefTitle",
    # Conditions (used for indication matching)
    "protocolSection.conditionsModule.conditions",
    # Interventions + sponsor (used for asset matching)
    "protocolSection.armsInterventionsModule.interventions",
    "protocolSection.sponsorCollaboratorsModule.leadSponsor",
    "protocolSection.sponsorCollaboratorsModule.collaborators",
    # Status
    "protocolSection.statusModule.overallStatus",
    "protocolSection.statusModule.studyFirstPostDateStruct",
    "protocolSection.statusModule.lastUpdatePostDateStruct",
    # Design — useful for team manual review
    "protocolSection.designModule.phases",
    "protocolSection.designModule.studyType",
    "protocolSection.designModule.enrollmentInfo",
    # MeSH conditions — richer indication matching + manual review
    "derivedSection.conditionBrowseModule.meshes",
])


# ── Department paths (set by init_dept) ───────────────────────────────────────
DEPT_DIR:        Path = None
STATE_FILE:      Path = None
COMPARISON_FILE: Path = None
ALERT_CONFIG:    Path = None   # dept/alert_config.json
SMTP_CONFIG:     Path = None   # root/smtp_config.json
DEPT_NAME:       str  = None   # set by init_dept, used for all DB calls
INDICATION:      str  = ""     # "" = asset pipeline; "Ovarian Cancer" etc = indication pipeline


def init_dept(dept_arg: str):
    """
    Resolve department folder from CLI arg and set all global path variables.
    dept_arg can be:
      - A short name like "ADC"  -> resolved relative to this script's folder
      - A full path               -> used as-is
    """
    global DEPT_DIR, STATE_FILE, COMPARISON_FILE, ALERT_CONFIG, SMTP_CONFIG, DEPT_NAME

    _here     = Path(__file__).parent
    dept_path = Path(dept_arg) if Path(dept_arg).is_absolute() else _here / dept_arg

    dept_path.mkdir(parents=True, exist_ok=True)

    DEPT_DIR        = dept_path
    DEPT_NAME       = dept_path.name
    STATE_FILE      = dept_path / "state.json"
    COMPARISON_FILE = dept_path / "latest_comparison.xlsx"
    ALERT_CONFIG    = dept_path / "alert_config.json"
    SMTP_CONFIG     = _here / "smtp_config.json"

    print(f"  Department    : {dept_path.name}  ({dept_path})")


# ── State ──────────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {"etag": None, "last_run_date": None}


def save_state(etag: str, last_run_date: str):
    STATE_FILE.write_text(
        json.dumps({"etag": etag, "last_run_date": last_run_date}, indent=2),
        encoding="utf-8",
    )
    print(f"  State saved: last_run_date={last_run_date}")


# ── Keyword loading ────────────────────────────────────────────────────────────

# Column names recognised as generic keyword columns (case-insensitive)
_GENERIC_KW_COLS = {
    "keyword", "keywords", "term", "terms",
    "search word", "search words", "search term", "search terms", "word", "words",
}

# Column names that trigger Drug Name / Alias Name parsing
_DRUG_NAME_COL  = "drug name"
_ALIAS_NAME_COL = "alias name"


def _parse_drug_alias_row(drug_name: str | None, alias_name: str | None) -> list[str]:
    """
    Extract keywords from one Drug Name / Alias Name row.
      - Drug name is always added (if non-empty).
      - Alias name is pipe-split; each part is added (skip "-" or blank).
    """
    terms: list[str] = []

    if drug_name and str(drug_name).strip() not in ("", "-"):
        terms.append(str(drug_name).strip().lower())

    if alias_name and str(alias_name).strip() not in ("", "-"):
        for part in str(alias_name).split("|"):
            part = part.strip()
            if part and part != "-":
                terms.append(part.lower())

    return terms


def _load_txt(path: Path) -> list[str]:
    """Read keywords from a plain-text file — one per line, # = comment."""
    lines = path.read_text(encoding="utf-8").splitlines()
    return [l.strip().lower() for l in lines if l.strip() and not l.startswith("#")]


def _load_xlsx(path: Path) -> list[str]:
    """
    Read keywords from an Excel file.

    Auto-detects format:
      A) Drug Name + Alias Name columns  -> drug/alias parsing with pipe split
      B) Recognised keyword column       -> reads that column directly
      C) Fallback                        -> reads first column
    """
    try:
        wb      = load_workbook(str(path), read_only=True, data_only=True)
        ws      = wb.active
        headers = None
        terms: list[str] = []

        drug_col  = None
        alias_col = None
        kw_col    = None

        for row in ws.iter_rows(values_only=True):
            # First row: detect format
            if headers is None:
                headers = [str(h).strip().lower() if h else "" for h in row]

                # Check for Drug Name / Alias Name columns
                for i, h in enumerate(headers):
                    if h == _DRUG_NAME_COL:
                        drug_col = i
                    if h == _ALIAS_NAME_COL:
                        alias_col = i

                if drug_col is not None:
                    # Drug/alias format detected
                    continue

                # Otherwise look for a generic keyword column
                for i, h in enumerate(headers):
                    if h in _GENERIC_KW_COLS:
                        kw_col = i
                        break

                if kw_col is None:
                    kw_col = 0   # fallback: first column
                continue

            # Data rows
            if drug_col is not None:
                # Drug Name / Alias Name format
                drug  = row[drug_col]  if drug_col  < len(row) else None
                alias = row[alias_col] if alias_col is not None and alias_col < len(row) else None
                terms.extend(_parse_drug_alias_row(drug, alias))
            else:
                val = row[kw_col] if kw_col < len(row) else None
                if val and str(val).strip() and str(val).strip() != "-":
                    terms.append(str(val).strip().lower())

        wb.close()
        return terms

    except Exception as e:
        print(f"    WARNING: Could not read {path.name} ({e})")
        return []


def load_keywords() -> list[str]:
    """
    Load keywords for the current run:
      - Indication pipeline (INDICATION != ''): keywords from dept_indications table.
      - Asset pipeline     (INDICATION == ''): drug names/aliases + general terms.
    """
    try:
        if INDICATION:
            kws = db.load_indication_keywords(DEPT_NAME, INDICATION)
            print(f"  dept_indications     {len(kws):>5} terms  (indication: {INDICATION})")
            return kws

        seen:    set[str]  = set()
        all_kws: list[str] = []

        # Drug names + aliases from dept_keywords
        drug_kw_count = 0
        for row in db.load_dept_keywords(DEPT_NAME):
            terms = _parse_drug_alias_row(row.get("drug_name"), row.get("alias_names", ""))
            for t in terms:
                if t not in seen:
                    seen.add(t)
                    all_kws.append(t)
                    drug_kw_count += 1
        print(f"  dept_keywords        {drug_kw_count:>5} terms")

        # General terms from dept_general_terms
        gt_count = 0
        for term in db.load_dept_general_terms(DEPT_NAME):
            t = str(term or "").strip().lower()
            if t and t not in seen:
                seen.add(t)
                all_kws.append(t)
                gt_count += 1
        print(f"  dept_general_terms   {gt_count:>5} terms")

        print(f"  Total unique keywords: {len(all_kws)}")
        return all_kws

    except Exception as e:
        print(f"  WARNING: Could not load keywords from DB ({e}) — keyword filter disabled.")
        return []


def matches_keywords(trial: dict, keywords: list[str]) -> bool:
    """
    Asset keyword match — searches Title, Conditions, Interventions, Sponsor.
    Exact word-boundary, case-insensitive.
    """
    if not keywords:
        return False
    haystack = " | ".join([
        trial.get("Title",         ""),
        trial.get("Conditions",    ""),
        trial.get("Interventions", ""),
        trial.get("Sponsor",       ""),
    ]).lower()
    for kw in keywords:
        pattern = r'\b' + re.escape(kw.lower()) + r'\b'
        if re.search(pattern, haystack):
            return True
    return False


def matches_indication_keywords(trial: dict, keywords: list[str]) -> bool:
    """
    Indication keyword match — searches Title, Conditions, MeSH Conditions.
    Exact word-boundary, case-insensitive.
    Deliberately excludes Interventions and Sponsor (not relevant for disease matching).
    """
    if not keywords:
        return False
    haystack = " | ".join([
        trial.get("Title",           ""),
        trial.get("Conditions",      ""),
        trial.get("MeSH Conditions", ""),
    ]).lower()
    for kw in keywords:
        pattern = r'\b' + re.escape(kw.lower()) + r'\b'
        if re.search(pattern, haystack):
            return True
    return False


# ── ETag + date range ──────────────────────────────────────────────────────────

def check_etag(stored_etag: str | None) -> tuple[bool, str]:
    req_headers = dict(PUB_HEADERS)
    if stored_etag:
        req_headers["If-None-Match"] = stored_etag

    print(f"  GET {RSS_URL}")
    resp = requests.get(RSS_URL, headers=req_headers, timeout=30)

    if resp.status_code == 304:
        print("  HTTP 304 Not Modified — no new CT.gov batch.")
        return False, stored_etag

    resp.raise_for_status()
    new_etag = resp.headers.get("etag", "")
    print("  HTTP 200 — New batch detected.")
    return True, new_etag


def date_range(last_run_date: str | None) -> tuple[str, str]:
    today = datetime.today().strftime("%Y-%m-%d")
    if last_run_date:
        from_date = last_run_date  # re-scan the last run day to catch late CT.gov updates
    else:
        from_date = (datetime.today() - timedelta(days=10)).strftime("%Y-%m-%d")
    return from_date, today


# ── FLOW 1 — Global Discovery ──────────────────────────────────────────────────

def _parse_scan_trial(study: dict) -> dict:
    ps    = study.get("protocolSection", {})
    ds    = study.get("derivedSection", {})
    imod  = ps.get("identificationModule", {})
    smod  = ps.get("statusModule", {})
    cmod  = ps.get("conditionsModule", {})
    amod  = ps.get("armsInterventionsModule", {})
    spmod = ps.get("sponsorCollaboratorsModule", {})
    dmod  = ps.get("designModule", {})
    cbmod = ds.get("conditionBrowseModule", {})
    nct_id = imod.get("nctId", "")
    return {
        "NCT ID":             nct_id,
        "Title":              imod.get("briefTitle", ""),
        "Link":               f"https://clinicaltrials.gov/study/{nct_id}",
        "Conditions":         "; ".join(cmod.get("conditions", [])),
        "MeSH Conditions":    "; ".join(
            m.get("term", "") for m in cbmod.get("meshes", []) if m.get("term")
        ),
        "Interventions":      "; ".join(
            f"{i.get('type','')}: {i.get('name','')}"
            for i in amod.get("interventions", [])
        ),
        "Sponsor":            spmod.get("leadSponsor", {}).get("name", ""),
        "Collaborators":      "; ".join(
            c.get("name", "") for c in spmod.get("collaborators", []) if c.get("name")
        ),
        "Recruitment Status": smod.get("overallStatus", ""),
        "Phase":              "; ".join(dmod.get("phases", [])),
        "Study Type":         dmod.get("studyType", ""),
        "Enrollment":         str(dmod.get("enrollmentInfo", {}).get("count", "") or ""),
        "Study First Posted": smod.get("studyFirstPostDateStruct", {}).get("date", ""),
        "Last Updated":       smod.get("lastUpdatePostDateStruct", {}).get("date", ""),
    }


def scan_global_delta(from_date: str, to_date: str) -> list[dict]:
    """
    Paginated REST API scan — one pass (~9 pages) returns all ~8,575 changed NCTs
    with enough fields for classification, keyword matching, and Excel output.
    No individual per-NCT fetches needed for this step.
    """
    params     = {
        "filter.advanced": f"AREA[LastUpdatePostDate]RANGE[{from_date},{to_date}]",
        "fields":          SCAN_FIELDS,
        "pageSize":        1000,
        "format":          "json",
    }
    results    = []
    page       = 1
    page_token = None

    while True:
        if page_token:
            params["pageToken"] = page_token

        resp = requests.get(API_URL, headers=PUB_HEADERS, params=params, timeout=60)
        resp.raise_for_status()
        data    = resp.json()
        studies = data.get("studies", [])

        for s in studies:
            results.append(_parse_scan_trial(s))

        print(f"    Page {page}: {len(studies)} studies  (total: {len(results)})")

        next_token = data.get("nextPageToken")
        if not next_token:
            break
        page_token = next_token
        page += 1

    return results


def run_flow1(
    from_date:   str,
    to_date:     str,
    tracking_db: dict[str, str],
    keywords:    list[str],
    indication:  str = "",
) -> tuple[list[dict], list[dict], list[dict]]:
    """
    Returns:
        modified_trials  in tracking list, date changed
        new_candidates   not in list, keyword matched
        unmatched        not in list, no keyword match (for gap review)

    Uses matches_indication_keywords() when indication is set,
    matches_keywords() for the asset pipeline.
    """
    print(f"\n[FLOW 1] Scanning  {from_date} -> {to_date}")
    all_trials = scan_global_delta(from_date, to_date)
    print(f"  Total in delta: {len(all_trials)}")

    matcher = matches_indication_keywords if indication else matches_keywords

    modified_trials = []
    new_candidates  = []
    unmatched       = []
    stale_count     = 0

    for trial in all_trials:
        nct_id     = trial["NCT ID"]
        api_date   = trial["Last Updated"]
        known_date = tracking_db.get(nct_id)

        if known_date is not None:
            if api_date != str(known_date):
                modified_trials.append(trial)
            else:
                stale_count += 1
        else:
            if matcher(trial, keywords):
                new_candidates.append(trial)
            else:
                unmatched.append(trial)

    print(f"  Modified       : {len(modified_trials)}")
    print(f"  Stale (skip)   : {stale_count}")
    print(f"  New candidates : {len(new_candidates)}  (keyword match)")
    print(f"  Unmatched      : {len(unmatched)}  (no match — sent for gap review)")
    return modified_trials, new_candidates, unmatched


# ── FLOW 2 — Field-Level Diff ──────────────────────────────────────────────────

def load_prev_versions(path: Path) -> tuple[dict[str, int], dict[str, dict]]:
    """Load stored version numbers and row data from previous comparison Excel."""
    versions: dict[str, int]  = {}
    rows:     dict[str, dict] = {}
    if not path.exists():
        return versions, rows
    try:
        wb      = load_workbook(str(path), read_only=True, data_only=True)
        ws      = wb.active
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = list(row)
                continue
            r   = dict(zip(headers, row))
            nct = str(r.get("NCT ID", "") or "").strip()
            ver = r.get("Curr Version")
            if not nct:
                continue
            if ver is not None:
                try:
                    versions[nct] = int(ver)
                except (ValueError, TypeError):
                    pass
            rows[nct] = {
                "nct_id":             nct,
                "note":               str(r.get("Note", "") or ""),
                "total_versions":     r.get("Total Versions", ""),
                "prev_version":       r.get("Prev Version", ""),
                "curr_version":       r.get("Curr Version", ""),
                "prev_date":          r.get("Prev Date", ""),
                "curr_date":          r.get("Curr Date", ""),
                "curr_status":        str(r.get("Current Status", "") or ""),
                "modules_changed":    str(r.get("Modules Changed", "") or ""),
                "field_change_count": r.get("Field Change Count", ""),
                "field_changes":      str(r.get("Field Changes", "") or ""),
            }
        wb.close()
        print(f"  Loaded {len(versions)} stored versions from {path.name}")
    except Exception as e:
        print(f"  Could not load {path.name} ({e}), running fresh.")
    return versions, rows


def _compare_pair(nct_id: str, prev_meta: dict, curr_meta: dict, total_versions: int) -> dict:
    """Fetch and diff two specific CT.gov versions. Returns same schema as get_latest_comparison()."""
    module_labels = curr_meta.get("moduleLabels", [])
    prev_data = get_version_data(nct_id, prev_meta["version"])
    curr_data = get_version_data(nct_id, curr_meta["version"])
    field_chg = compare_versions(prev_data, curr_data, module_labels)
    return {
        "nct_id":             nct_id,
        "note":               "",
        "total_versions":     total_versions,
        "prev_version":       prev_meta["version"],
        "curr_version":       curr_meta["version"],
        "prev_date":          prev_meta.get("date", ""),
        "curr_date":          curr_meta.get("date", ""),
        "curr_status":        curr_meta.get("status", ""),
        "modules_changed":    "; ".join(module_labels),
        "field_change_count": len(field_chg),
        "field_changes":      format_field_changes(field_chg),
        "prev_full_data":     prev_data,
        "curr_full_data":     curr_data,
    }


def run_flow2(
    nct_ids:        list[str],
    prev_versions:  dict[str, int],
    prev_rows:      dict[str, dict],
    drug_terms:     list = (),
    base_organized: dict = None,
) -> tuple[list[dict], list[dict], list[dict]]:
    """
    Returns:
        all_rows        one row per NCT (diffs + carried-forward)
        newly_compared  only NCTs with a new version this run
        redirects       list of {requested, canonical, in_input}

    Changed NCTs are parsed and written to organized_trials + nct_version_pairs
    incrementally every 50 new versions — no large in-memory accumulation.
    """
    print(f"\n[FLOW 2] Checking {len(nct_ids)} tracked NCTs...")

    total               = len(nct_ids)
    all_rows            = []
    newly_compared      = []
    redirects           = []
    version_pairs_batch = []
    version_pairs_saved = 0
    fresh_org_batch     = []
    fresh_org_saved     = 0
    input_id_set        = set(nct_ids)
    skipped             = 0
    consecutive_errors  = 0

    for i, nct_id in enumerate(nct_ids, 1):
        try:
            changes    = get_history_list(nct_id)
            api_latest = changes[-1]["version"] if changes else None
            stored_ver = prev_versions.get(nct_id)
            consecutive_errors = 0  # reset on any successful API call

            # No version history on CT.gov — save placeholder, skip on future runs
            if api_latest is None:
                placeholder = {
                    "nct_id": nct_id, "note": "no version history",
                    "total_versions": 0, "prev_version": 0, "curr_version": 0,
                    "prev_date": "", "curr_date": "", "curr_status": "",
                    "modules_changed": "", "field_change_count": 0, "field_changes": "",
                }
                all_rows.append(placeholder)
                if stored_ver is None:
                    newly_compared.append(placeholder)
                skipped += 1
                print(f"  [{i}/{total}] {nct_id} — no version history, skipped")
                time.sleep(0.1)
                continue

            # Same version — carry forward without any additional API calls
            if api_latest is not None and stored_ver is not None and api_latest == stored_ver:
                carried = prev_rows.get(nct_id, {
                    "nct_id": nct_id, "note": "", "total_versions": len(changes),
                    "prev_version": "", "curr_version": api_latest,
                    "prev_date": "", "curr_date": changes[-1].get("date", ""),
                    "curr_status": changes[-1].get("status", ""),
                    "modules_changed": "", "field_change_count": 0, "field_changes": "",
                })
                all_rows.append(carried)
                skipped += 1
                print(f"  [{i}/{total}] {nct_id} — v{api_latest} unchanged, skipped")
                time.sleep(0.1)
                continue

            # New version(s) — gap-fill all intermediate versions
            total_versions = len(changes)

            if stored_ver is None:
                # First encounter — compare latest pair only (no stored baseline to chain from)
                print(f"  [{i}/{total}] {nct_id} — first diff, comparing latest pair...")
                pair_rows = [get_latest_comparison(nct_id)]
            else:
                new_metas = [c for c in changes if c["version"] > stored_ver]
                prev_meta  = next((c for c in changes if c["version"] == stored_ver), None)

                if prev_meta is None:
                    # stored_ver dropped from CT.gov history (very rare) — fall back
                    print(f"  [{i}/{total}] {nct_id} — stored v{stored_ver} not in history, latest diff only...")
                    pair_rows = [get_latest_comparison(nct_id)]
                else:
                    n_gaps = len(new_metas)
                    print(f"  [{i}/{total}] {nct_id} — {n_gaps} new version(s) since v{stored_ver}, gap-filling...")
                    chain     = [prev_meta] + new_metas
                    pair_rows = []
                    for j in range(len(chain) - 1):
                        pair_rows.append(_compare_pair(nct_id, chain[j], chain[j + 1], total_versions))
                        if j < len(chain) - 2:
                            time.sleep(0.3)

            # Latest row → version_cache (all_rows);  all rows → field_changes_log (newly_compared)
            row = pair_rows[-1]
            all_rows.append(row)
            newly_compared.extend(pair_rows)

            # Fetch full data: parse organized + canonical check + version pairs
            data, _ = fetch_study_data(nct_id)
            if data:
                raw_row      = build_raw_row(data)
                canonical_id = raw_row.get("NCT ID") or nct_id

                # Parse immediately — no large list accumulation
                parsed = parse_study_to_organized(nct_id, data)
                if base_organized is not None:
                    base_organized[nct_id] = parsed
                fresh_org_batch.append(parsed)

                if canonical_id != nct_id:
                    already_in = canonical_id in input_id_set
                    redirects.append({
                        "requested": nct_id,
                        "canonical": canonical_id,
                        "in_input":  already_in,
                    })
                    label = "already tracked" if already_in else "NOT in input — will replace"
                    print(f"    REDIRECT: {nct_id} -> {canonical_id} ({label})")

                # Store ALL intermediate pairs in nct_version_pairs
                for pr in pair_rows:
                    version_pairs_batch.append({
                        "nct_id":             nct_id,
                        "note":               pr.get("note", ""),
                        "total_versions":     pr.get("total_versions", 0),
                        "prev_version":       pr.get("prev_version"),
                        "curr_version":       pr.get("curr_version"),
                        "prev_date":          pr.get("prev_date", ""),
                        "curr_date":          pr.get("curr_date", ""),
                        "curr_status":        pr.get("curr_status", ""),
                        "modules_changed":    pr.get("modules_changed", ""),
                        "field_change_count": pr.get("field_change_count", 0),
                        "field_changes":      pr.get("field_changes", ""),
                        "prev_full_data":     pr.get("prev_full_data"),
                        "curr_full_data":     pr.get("curr_full_data"),
                    })

                # Flush every 50 — keeps memory flat regardless of total NCT count
                if len(version_pairs_batch) >= 50:
                    db.upsert_version_pairs(DEPT_NAME, version_pairs_batch)
                    version_pairs_saved += len(version_pairs_batch)
                    print(f"  -- Flushed version pairs ({version_pairs_saved} total) --")
                    version_pairs_batch = []

                if len(fresh_org_batch) >= 50:
                    db.upsert_organized_trials(DEPT_NAME, fresh_org_batch)
                    fresh_org_saved += len(fresh_org_batch)
                    print(f"  -- Flushed organized trials ({fresh_org_saved} total) --")
                    fresh_org_batch = []

            time.sleep(0.5)

        except Exception as e:
            err = str(e)
            print(f"  [{i}/{total}] {nct_id} ERROR: {err}")
            all_rows.append({
                "nct_id": nct_id, "note": f"ERROR: {err}",
                "total_versions": "", "prev_version": "", "curr_version": "",
                "prev_date": "", "curr_date": "", "curr_status": "",
                "modules_changed": "", "field_change_count": "", "field_changes": "",
            })
            if any(code in err for code in ("429", "503", "Too Many", "rate limit")):
                consecutive_errors += 1
                if consecutive_errors >= 3:
                    print(f"  [FLOW 2] CT.gov rate-limiting — sleeping 60s before continuing...")
                    time.sleep(60)
                    consecutive_errors = 0
                else:
                    time.sleep(5)
            else:
                consecutive_errors = 0
                time.sleep(0.2)

        # Periodic reconnect every 500 NCTs — keeps connection age short on multi-hour runs
        if i % 500 == 0:
            db._conn = None
            print(f"  -- [{i}/{total}] Periodic DB reconnect --")

    # Flush any remaining items
    if version_pairs_batch:
        db.upsert_version_pairs(DEPT_NAME, version_pairs_batch)
        version_pairs_saved += len(version_pairs_batch)

    if fresh_org_batch:
        db.upsert_organized_trials(DEPT_NAME, fresh_org_batch)
        fresh_org_saved += len(fresh_org_batch)

    print(f"\n[FLOW 2] Complete.")
    print(f"  Skipped (no change)  : {skipped}")
    print(f"  New diffs            : {len(newly_compared)}")
    print(f"  Redirects            : {len(redirects)}")
    print(f"  Organized updated    : {fresh_org_saved}")
    print(f"  Version pairs saved  : {version_pairs_saved}")
    return all_rows, newly_compared, redirects


# ── Canonical auto-update ──────────────────────────────────────────────────────

def apply_canonical_updates(redirects: list[dict], nct_ids: list[str]) -> list[str]:
    """Replace redirected IDs in the in-memory tracking list. Supabase is updated by the caller."""
    if not redirects:
        print("  No canonical updates to apply.")
        return nct_ids

    redirect_map = {r["requested"]: r for r in redirects}
    seen:     set[str]    = set()
    cleaned:  list[str]   = []
    replaced: list[tuple] = []
    removed:  list[tuple] = []

    for nct in nct_ids:
        if nct in redirect_map:
            r         = redirect_map[nct]
            canonical = r["canonical"]
            if r["in_input"]:
                removed.append((nct, f"redirects to {canonical} (already tracked)"))
            elif canonical not in seen:
                cleaned.append(canonical)
                seen.add(canonical)
                replaced.append((nct, canonical))
            else:
                removed.append((nct, f"redirects to {canonical} (already added)"))
        elif nct in seen:
            removed.append((nct, "duplicate"))
        else:
            cleaned.append(nct)
            seen.add(nct)

    if replaced:
        print(f"  Replaced {len(replaced)} canonical ID(s):")
        for old, new in replaced:
            print(f"    {old} -> {new}")
    if removed:
        print(f"  Removed {len(removed)} duplicate/redirect ID(s).")
    print(f"  Tracking list updated: {len(cleaned)} IDs.")
    return cleaned


def _parse_raw_into_base(raw_api_jsons: list[dict], base: dict):
    """Parse raw CT.gov API JSON dicts into 62-column organized rows and upsert into base."""
    if not raw_api_jsons:
        return
    for data in raw_api_jsons:
        nct_id = (
            data.get("protocolSection", {})
            .get("identificationModule", {})
            .get("nctId", "")
        )
        if nct_id:
            base[nct_id] = parse_study_to_organized(nct_id, data)


def sync_baseline(nct_ids: list[str], drug_terms: list = (), batch_size: int = 50) -> dict:
    """
    Step 0: For every tracked NCT:
      - Pass 1: Ensure its current organized data is in organized_trials.
      - Pass 2: Backfill all version pair comparisons into nct_version_pairs.

    Pass 1 and Pass 2 are fully independent — each checks its own table and
    skips already-saved NCTs. A crash in either pass loses at most batch_size
    NCTs of work. Re-running picks up exactly where it left off.

    Returns the {nct_id: row_dict} base dict for reuse in Step 5.
    """
    # ── Pass 1: Organized data ────────────────────────────────────────────────
    base    = db.load_organized_base(DEPT_NAME)
    missing = [n for n in nct_ids if n not in base]

    if not missing:
        print(f"  [Pass 1] All {len(nct_ids)} NCTs already in organized_trials — skipping.")
    else:
        total     = len(missing)
        raw_batch = []
        saved_org = 0
        print(f"\n  [Pass 1] Fetching organized data for {total} missing NCTs (saving every {batch_size})...")

        for i, nct_id in enumerate(missing, 1):
            try:
                data, _ = fetch_study_data(nct_id)
                if data:
                    raw_batch.append(data)
                    print(f"  [{i}/{total}] {nct_id} — fetched")
                else:
                    print(f"  [{i}/{total}] {nct_id} — no data returned")
            except Exception as e:
                print(f"  [{i}/{total}] {nct_id} — WARNING: {e}")
            time.sleep(0.3)

            if len(raw_batch) >= batch_size:
                _parse_raw_into_base(raw_batch, base)
                db.upsert_organized_trials(DEPT_NAME, list(base.values()))
                saved_org += len(raw_batch)
                print(f"  -- Saved organized batch ({saved_org} total) --")
                raw_batch = []

        if raw_batch:
            _parse_raw_into_base(raw_batch, base)
            db.upsert_organized_trials(DEPT_NAME, list(base.values()))
            saved_org += len(raw_batch)
            print(f"  -- Saved final organized batch ({saved_org} total) --")

        print(f"  [Pass 1] Complete — {saved_org} NCTs saved to organized_trials")

    # ── Pass 2: Version pair history (independent of Pass 1) ─────────────────
    already_vp = db.load_ncts_with_version_pairs(DEPT_NAME)
    vp_missing = [n for n in nct_ids if n not in already_vp]

    if not vp_missing:
        print(f"  [Pass 2] All {len(nct_ids)} NCTs already have version pair history — skipping.")
    else:
        print(f"\n  [Pass 2] Backfilling version pair history for {len(vp_missing)} NCTs (saving every {batch_size})...")
        try:
            _backfill_version_pairs(vp_missing, batch_size)
        except Exception as e:
            print(f"  [Pass 2] WARNING: backfill aborted — {e}")
            print(f"  [Pass 2] Pipeline will continue. Missing NCTs will be retried on next run.")

    return base


def _backfill_version_pairs(nct_ids: list[str], batch_size: int = 50) -> None:
    """Fetch and save all version pair history for given NCTs, saving every batch_size NCTs."""
    total           = len(nct_ids)
    vp_batch        = []
    saved_vp        = 0
    consecutive_403 = 0

    for i, nct_id in enumerate(nct_ids, 1):
        try:
            pairs = _get_all_version_pairs(nct_id)
            consecutive_403 = 0  # reset on any success (even empty = no history)
            if pairs:
                for p in pairs:
                    p["nct_id"] = nct_id
                vp_batch.extend(pairs)
            else:
                # Sentinel so NCT is not retried every run
                vp_batch.append({
                    "nct_id": nct_id, "note": "no version history",
                    "total_versions": 0, "prev_version": 0, "curr_version": 0,
                    "prev_date": "", "curr_date": "", "curr_status": "",
                    "modules_changed": "", "field_change_count": 0, "field_changes": "",
                })
            print(f"  [{i}/{total}] {nct_id} — {len(pairs)} pair(s) queued")
        except Exception as e:
            err = str(e)
            if "403" in err or "429" in err or "503" in err:
                consecutive_403 += 1
                print(f"  [{i}/{total}] {nct_id} — WARNING: {e}")
                if consecutive_403 >= 3:
                    print(f"  [Pass 2] CT.gov is rate-limiting (3 consecutive 403s) — aborting backfill.")
                    print(f"  [Pass 2] {total - i} NCTs skipped; will retry on next pipeline run.")
                    break
            else:
                consecutive_403 = 0
                print(f"  [{i}/{total}] {nct_id} — WARNING: {e}")
        time.sleep(1.5)

        if len(vp_batch) >= batch_size:
            db.upsert_version_pairs(DEPT_NAME, vp_batch)
            saved_vp += len(vp_batch)
            print(f"  -- Saved version pair batch ({saved_vp} total pairs) --")
            vp_batch = []

    if vp_batch:
        db.upsert_version_pairs(DEPT_NAME, vp_batch)
        saved_vp += len(vp_batch)
        print(f"  -- Saved final version pair batch ({saved_vp} total pairs) --")

    print(f"  Pass 2 complete — {saved_vp} version pairs saved to nct_version_pairs")




# ── Alert ─────────────────────────────────────────────────────────────────────

def send_alert(
    run_date:        str,
    new_candidates:  list[dict],
    unmatched:       list[dict],
    modified_trials: list[dict],
    newly_compared:  list[dict],
    redirects:       list[dict],
):
    """
    Send a department-specific email alert with dashboard link and statistics.
    Silently skips if alert_config.json or smtp_config.json is missing/disabled.
    """
    # Load dept alert config
    if not ALERT_CONFIG.exists():
        print("  No alert_config.json found — skipping email.")
        return
    try:
        acfg = json.loads(ALERT_CONFIG.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  Could not read alert_config.json ({e}) — skipping email.")
        return

    if not acfg.get("enabled", True):
        print("  Alert disabled in alert_config.json — skipping email.")
        return

    to_list       = acfg.get("to",  [])
    cc_list       = acfg.get("cc",  [])
    dept_name     = acfg.get("dept_name", DEPT_DIR.name)
    dashboard_url = acfg.get("dashboard_url", f"https://nct-dashboard-tan.vercel.app/?dept={dept_name}")

    if not to_list:
        print("  No recipients in alert_config.json — skipping email.")
        return

    # Load SMTP config
    if not SMTP_CONFIG.exists():
        print("  No smtp_config.json found — skipping email.")
        return
    try:
        scfg = json.loads(SMTP_CONFIG.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  Could not read smtp_config.json ({e}) — skipping email.")
        return

    smtp_host = scfg.get("smtp_host", "")
    smtp_port = int(scfg.get("smtp_port", 587))
    smtp_user = scfg.get("smtp_user", "")
    smtp_pass = scfg.get("smtp_password", "")
    from_addr = scfg.get("from_email", smtp_user)
    from_name = scfg.get("from_name", "NCT Tracker")

    if not smtp_host or not smtp_user:
        print("  smtp_config.json incomplete — skipping email.")
        return

    # Subject
    subject = (
        f"[{dept_name} Tracker] CT.gov — {run_date} | "
        f"{len(new_candidates)} New | {len(modified_trials)} Modified"
    )

    # Action items
    action_rows = []
    if new_candidates:
        action_rows.append((
            f"New {dept_name} Trials Found",
            f"{len(new_candidates)} keyword-matched trials require review. "
            f"Visit the dashboard to review and approve trials relevant to your pipeline."
        ))
    if unmatched:
        action_rows.append((
            "Unmatched Trials - Gap Review",
            f"{len(unmatched)} trials did not match any keyword. "
            f"Review these on the dashboard for any {dept_name}-relevant trials that were missed, "
            "then update keywords to capture them going forward."
        ))
    if newly_compared:
        action_rows.append((
            "Field-Level Changes",
            f"{len(newly_compared)} tracked trials have a new version this week. "
            "Review the field-level diff on the dashboard to understand what changed "
            "(status, enrollment, dates, eligibility, etc.)."
        ))
    if redirects:
        action_rows.append((
            "Redirected NCT IDs",
            f"{len(redirects)} NCT IDs were redirected to a canonical ID. "
            "The tracking list has been auto-updated — no action needed."
        ))

    action_html = ""
    for i, (sheet, desc) in enumerate(action_rows, 1):
        action_html += f"""
        <tr>
          <td style="padding:10px 12px;vertical-align:top;width:30px;color:#b07d00;font-weight:bold;font-size:14px;">{i}.</td>
          <td style="padding:10px 12px;vertical-align:top;">
            <span style="font-weight:bold;color:#1a3a5c;">{sheet}</span><br>
            <span style="color:#555;font-size:13px;">{desc}</span>
          </td>
        </tr>"""

    if not action_rows:
        action_html = """
        <tr><td colspan="2" style="padding:10px 12px;color:#555;font-size:13px;">
          No new candidates or changes detected this run. No action required.
        </td></tr>"""

    # Summary rows
    summary_data = [
        (f"New {dept_name} Trials Found",  len(new_candidates),  "#C0392B", "Keyword matched — pending team review"),
        ("Unmatched Trials - Gap Review",  len(unmatched),       "#7B68EE", "No keyword match — review for gaps"),
        ("CT.gov Updated This Week",       len(modified_trials), "#D35400", "Tracked trials flagged as updated"),
        ("Field-Level Changes",            len(newly_compared),  "#117A65", "New versions detected in tracked trials"),
        ("Redirected NCT IDs",             len(redirects),       "#555555", "Canonical ID fixes auto-applied"),
    ]
    summary_html = ""
    for sheet, count, color, desc in summary_data:
        summary_html += f"""
        <tr>
          <td style="padding:9px 14px;font-size:13px;color:#333;border-bottom:1px solid #eee;">{sheet}</td>
          <td style="padding:9px 14px;font-size:15px;font-weight:bold;color:{color};text-align:center;border-bottom:1px solid #eee;">{count:,}</td>
          <td style="padding:9px 14px;font-size:12px;color:#777;border-bottom:1px solid #eee;">{desc}</td>
        </tr>"""

    # Logo path
    logo_path = Path(__file__).parent / "Phenomiqs_Logo.png"
    has_logo  = logo_path.exists()
    logo_tag  = '<img src="cid:phenomiqs_logo" alt="Phenomiqs" style="height:52px;" />' if has_logo else \
                '<span style="color:#0d2d5e;font-size:20px;font-weight:bold;letter-spacing:1px;">Phenomiqs</span>'

    html_body = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background-color:#f0f2f5;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background-color:#f0f2f5;padding:24px 0;">
<tr><td align="center">
<table width="640" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.10);">

  <!-- Header / Logo — white background so logo is always visible -->
  <tr>
    <td style="background-color:#ffffff;padding:18px 32px;text-align:left;border-bottom:3px solid #0d2d5e;">
      {logo_tag}
    </td>
    <td style="background-color:#ffffff;padding:18px 32px;text-align:right;border-bottom:3px solid #0d2d5e;">
      <span style="color:#0d2d5e;font-size:12px;font-weight:600;letter-spacing:0.5px;">Automated Alert</span>
    </td>
  </tr>

  <!-- Banner — solid colour (gradients unsupported in Outlook) -->
  <tr>
    <td colspan="2" style="background-color:#0d2d5e;padding:20px 32px;">
      <div style="color:#ffffff;font-size:19px;font-weight:bold;">
        {dept_name} NCT Tracking Update
      </div>
    </td>
  </tr>

  <!-- Greeting -->
  <tr>
    <td colspan="2" style="padding:24px 32px 8px;">
      <p style="margin:0;font-size:15px;color:#222;">Hi Team,</p>
      <p style="margin:12px 0 0;font-size:13px;color:#555;line-height:1.6;">
        The CT.gov batch scan for <strong>{run_date}</strong> has completed.
        Please review the summary below and visit the dashboard to take the required actions.
      </p>
    </td>
  </tr>

  <!-- Run Summary Table -->
  <tr>
    <td colspan="2" style="padding:16px 32px 8px;">
      <div style="font-size:13px;font-weight:bold;color:#0d2d5e;text-transform:uppercase;
                  letter-spacing:0.8px;border-bottom:2px solid #0d2d5e;padding-bottom:6px;margin-bottom:0;">
        Run Summary
      </div>
      <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">
        <tr style="background-color:#0d2d5e;">
          <th style="padding:9px 14px;text-align:left;font-size:12px;color:#fff;font-weight:600;">Category</th>
          <th style="padding:9px 14px;text-align:center;font-size:12px;color:#fff;font-weight:600;width:70px;">Count</th>
          <th style="padding:9px 14px;text-align:left;font-size:12px;color:#fff;font-weight:600;">Description</th>
        </tr>
        {summary_html}
      </table>
    </td>
  </tr>

  <!-- Action Required -->
  <tr>
    <td colspan="2" style="padding:20px 32px 8px;">
      <div style="background-color:#fffbf0;border-left:4px solid #e8a020;border-radius:4px;padding:16px 18px;">
        <div style="font-size:13px;font-weight:bold;color:#b07d00;text-transform:uppercase;
                    letter-spacing:0.8px;margin-bottom:10px;">Action Required</div>
        <table width="100%" cellpadding="0" cellspacing="0">
          {action_html}
        </table>
      </div>
    </td>
  </tr>

  <!-- Dashboard link -->
  <tr>
    <td colspan="2" style="padding:16px 32px 24px;text-align:center;">
      <a href="{dashboard_url}"
         style="display:inline-block;background-color:#0d2d5e;color:#ffffff;font-size:14px;
                font-weight:bold;padding:12px 32px;border-radius:6px;text-decoration:none;
                letter-spacing:0.4px;">
        Open {dept_name} Dashboard &rarr;
      </a>
      <p style="margin:10px 0 0;font-size:11px;color:#aaa;">{dashboard_url}</p>
    </td>
  </tr>

  <!-- Footer -->
  <tr>
    <td colspan="2" style="background-color:#f5f7fa;padding:14px 32px;border-top:1px solid #e0e4ea;text-align:center;">
      <span style="font-size:11px;color:#aaa;">
        Phenomiqs NCT Tracker &nbsp;|&nbsp; Automated Alert &nbsp;|&nbsp; {run_date}
        &nbsp;&nbsp;&mdash;&nbsp;&nbsp;
        Do not reply to this email.
      </span>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>"""

    plain_fallback = (
        f"{dept_name} NCT Tracking Pipeline — {run_date}\n\n"
        f"New {dept_name} Trials Found : {len(new_candidates)}\n"
        f"Unmatched Trials             : {len(unmatched)}\n"
        f"CT.gov Updated               : {len(modified_trials)}\n"
        f"Field-Level Changes          : {len(newly_compared)}\n"
        f"Redirected NCT IDs           : {len(redirects)}\n\n"
        f"Dashboard: {dashboard_url}"
    )

    # Assemble MIME: mixed > related > alternative + logo | excel attachment
    msg_outer = MIMEMultipart("mixed")
    msg_outer["From"]    = f"{from_name} <{from_addr}>"
    msg_outer["To"]      = ", ".join(to_list)
    if cc_list:
        msg_outer["Cc"] = ", ".join(cc_list)
    msg_outer["Subject"] = subject

    msg_related = MIMEMultipart("related")
    msg_alt     = MIMEMultipart("alternative")
    msg_alt.attach(MIMEText(plain_fallback, "plain"))
    msg_alt.attach(MIMEText(html_body,      "html"))
    msg_related.attach(msg_alt)

    if has_logo:
        with open(logo_path, "rb") as f:
            logo_img = MIMEImage(f.read(), _subtype="png")
        logo_img.add_header("Content-ID",          "<phenomiqs_logo>")
        logo_img.add_header("Content-Disposition", "inline", filename="Phenomiqs_Logo.png")
        msg_related.attach(logo_img)

    msg_outer.attach(msg_related)

    # Send
    all_recipients = to_list + cc_list
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(from_addr, all_recipients, msg_outer.as_string())
        print(f"  Alert sent to: {', '.join(to_list)}"
              + (f"  CC: {', '.join(cc_list)}" if cc_list else ""))
    except Exception as e:
        print(f"  WARNING: Could not send email ({e})")


# ── Main ───────────────────────────────────────────────────────────────────────

def _run_unified(run_start: datetime) -> None:
    """
    Unified pipeline: fetch CT.gov delta ONCE for the dept, then classify
    each NCT against the asset pipeline and every indication pipeline
    simultaneously. Saves results per (dept, indication).

    This avoids repeated CT.gov fetches when running multiple indications —
    a single scan serves all pipelines.
    """
    print("  Mode: Unified (asset + all indications — single CT.gov fetch)")

    # ── Drug keywords (asset only) ─────────────────────────────────────────
    print("\n[Keywords]")
    global INDICATION
    INDICATION = ""           # ensure asset keyword loader is used
    asset_keywords  = load_keywords()
    drug_terms      = load_drug_keywords()
    drug_compiled   = _compile_drug_patterns(drug_terms)

    # ── Build pipeline registry ────────────────────────────────────────────
    # Each entry: indication → {tracking_set, rejected_set, keywords, matcher, new, modified, unmatched}
    all_indications = db.load_all_indications(DEPT_NAME)
    pipeline_keys   = [""] + all_indications    # "" = asset

    pipelines: dict[str, dict] = {}
    for ind in pipeline_keys:
        db.sync_approved_to_tracking(DEPT_NAME, ind)
        tracking_ids = db.load_tracking_list(DEPT_NAME, ind)
        rejected_ids = db.load_rejected_trials(DEPT_NAME, ind)

        if ind == "":
            kws     = asset_keywords
            matcher = matches_keywords
        else:
            kws     = db.load_indication_keywords(DEPT_NAME, ind)
            matcher = matches_indication_keywords

        label = ind or "Asset"
        print(f"  [{label:<25}]  {len(tracking_ids):>5} tracked  {len(kws):>5} keywords")

        pipelines[ind] = {
            "tracking_ids": tracking_ids,
            "tracking_set": set(tracking_ids),
            "rejected_set": set(rejected_ids),
            "keywords":     kws,
            "matcher":      matcher,
            "new":          [],
            "modified":     [],
            "unmatched":    [],
        }

    # ── Step 0 — Baseline sync (all pipelines share organized_trials) ──────
    all_tracked = list({nct for p in pipelines.values() for nct in p["tracking_ids"]})
    print(f"\n[Step 0] Baseline sync — {len(all_tracked)} unique tracked NCTs across all pipelines...")
    base_organized = sync_baseline(all_tracked, drug_compiled)

    # ── Step 1 — Load state ───────────────────────────────────────────────
    state         = db.load_pipeline_state(DEPT_NAME, "")
    new_etag      = state.get("etag") or ""
    last_run_date = state.get("last_run_date")
    today         = datetime.today().strftime("%Y-%m-%d")
    print(f"\n[State]  last_run={last_run_date or 'never'}")

    # ── Step 2 — FLOW 1: CT.gov date-range scan ───────────────────────────
    from_date, to_date = date_range(last_run_date)
    tracking_date_map = {
        nct: str(row.get("Last Update Post Date", "") or "")
        for nct, row in base_organized.items()
        if str(row.get("Last Update Post Date", "") or "").strip()
    }
    print(f"  Tracking date map: {len(tracking_date_map)} NCTs with stored dates")
    print(f"\n[Step 2] FLOW 1 — Single CT.gov fetch: {from_date} → {to_date}")
    all_trials = scan_global_delta(from_date, to_date)
    print(f"  Total in delta: {len(all_trials)}")

    # ── Step 3 — Classify each NCT against ALL pipelines in one pass ──────
    print(f"\n[Step 3] Classifying {len(all_trials)} trials across {len(pipeline_keys)} pipelines...")
    stale_counts: dict[str, int] = {ind: 0 for ind in pipeline_keys}

    for trial in all_trials:
        nct_id   = trial["NCT ID"]
        api_date = trial["Last Updated"]

        for ind, p in pipelines.items():
            if nct_id in p["rejected_set"]:
                continue
            if nct_id in p["tracking_set"]:
                known = tracking_date_map.get(nct_id)
                if known and api_date != known:
                    p["modified"].append(trial)
                else:
                    stale_counts[ind] += 1
            else:
                if p["matcher"](trial, p["keywords"]):
                    p["new"].append(trial)
                else:
                    p["unmatched"].append(trial)

    # Classification summary
    print(f"\n  {'Pipeline':<28} {'New':>6} {'Modified':>10} {'Unmatched':>11} {'Stale':>7}")
    print(f"  {'-'*66}")
    for ind, p in pipelines.items():
        label = ind or "Asset"
        print(f"  {label:<28} {len(p['new']):>6} {len(p['modified']):>10} "
              f"{len(p['unmatched']):>11} {stale_counts[ind]:>7}")

    # ── Step 4 — Flow 2: field-level diff (runs ONCE on all unique tracked NCTs) ──
    print(f"\n[Step 4] FLOW 2 — Field-Level Diff ({len(all_tracked)} unique tracked NCTs)...")
    prev_versions, prev_rows = db.load_version_cache(DEPT_NAME)
    all_diff_rows, newly_compared, redirects = run_flow2(
        all_tracked, prev_versions, prev_rows, drug_compiled, base_organized
    )

    # ── Step 5 — Canonical auto-update (asset tracking list) ──────────────
    print("\n[Step 5] Canonical ID auto-update...")
    asset_ids = apply_canonical_updates(redirects, pipelines[""]["tracking_ids"])
    if redirects:
        cleaned_set = set(asset_ids)
        for r in redirects:
            if r["requested"] not in cleaned_set:
                try:
                    db.delete_from_tracking_list(r["requested"], DEPT_NAME, "")
                except Exception:
                    pass
            if not r.get("in_input"):
                db.upsert_tracking_list(DEPT_NAME, [r["canonical"]], added_by="canonical", indication="")

    # ── Step 6 — Save per-pipeline results ────────────────────────────────
    print("\n[Step 6] Saving to DB...")
    new_unmatched_counts: dict[str, int] = {}
    for ind, p in pipelines.items():
        db.insert_new_candidates(DEPT_NAME, to_date, p["new"],      indication=ind)
        new_unmatched_counts[ind] = db.upsert_unmatched(DEPT_NAME, to_date, p["unmatched"], indication=ind) or 0
        db.insert_modified_log(  DEPT_NAME, to_date, p["modified"],  indication=ind)

    # Field changes and version cache are shared across all pipelines
    db.insert_field_changes(DEPT_NAME, to_date, newly_compared, indication="")
    # Also log field changes per indication so indication-specific views have data
    if newly_compared:
        for ind in all_indications:
            ind_tracking = pipelines[ind]["tracking_set"]
            ind_rows = [r for r in newly_compared if r["nct_id"] in ind_tracking]
            if ind_rows:
                db.insert_field_changes(DEPT_NAME, to_date, ind_rows, indication=ind)
    db.upsert_version_cache(DEPT_NAME, all_diff_rows)
    if redirects:
        db.insert_canonical_changes(DEPT_NAME, to_date, redirects)

    # ── Step 6c — Drug tagging (asset pipeline only) ──────────────────────
    print("\n[Step 6c] Tagging Primary Drug for all organized trials...")
    tag_all_organized_drug(DEPT_NAME, drug_terms)

    _duration = (datetime.now() - run_start).total_seconds()

    # ── Step 7 — Save state + run history for each pipeline ───────────────
    print("\n[Step 7] Saving state and run history...")
    for ind, p in pipelines.items():
        db.save_pipeline_state(DEPT_NAME, new_etag, to_date, indication=ind)
        db.insert_run_history(
            dept            = DEPT_NAME,
            run_date        = today,
            new_candidates  = len(p["new"]),
            unmatched       = new_unmatched_counts.get(ind, 0),  # only truly new unmatched
            modified_trials = len(p["modified"]),
            field_diffs     = len(newly_compared),
            canonical_fixes = len(redirects),
            duration_s      = _duration,
            output_file     = None,
            indication      = ind,
        )

    # ── Step 8 — Alert (asset pipeline summary) ───────────────────────────
    print("\n[Step 8] Sending alert...")
    send_alert(
        run_date        = to_date,
        new_candidates  = pipelines[""]["new"],
        unmatched       = pipelines[""]["unmatched"],
        modified_trials = pipelines[""]["modified"],
        newly_compared  = newly_compared,
        redirects       = redirects,
    )

    # ── Summary ───────────────────────────────────────────────────────────
    print("\n" + "=" * 66)
    print(f"UNIFIED RUN COMPLETE  [{DEPT_NAME}]  ({_duration:.0f}s)")
    print("=" * 66)
    print(f"  {'Pipeline':<28} {'New':>6} {'Modified':>10} {'Unmatched':>11}")
    print(f"  {'-'*58}")
    for ind, p in pipelines.items():
        label = ind or "Asset"
        print(f"  {label:<28} {len(p['new']):>6} {len(p['modified']):>10} {len(p['unmatched']):>11}")
    print(f"\n  Field Diffs (shared): {len(newly_compared)}   Canonical Fixes: {len(redirects)}")


def _run_single(run_start: datetime) -> None:
    """
    Single-pipeline mode (--indication X or asset only).
    Used for debugging / re-running one specific pipeline.
    Makes its own CT.gov fetch for the date range.
    """
    print(f"  Mode: Single — {INDICATION or 'Asset'}")

    print("\n[Keywords]")
    keywords      = load_keywords()
    drug_terms    = load_drug_keywords()
    drug_compiled = _compile_drug_patterns(drug_terms)

    print("\n[Approvals] Syncing dashboard approvals...")
    db.sync_approved_to_tracking(DEPT_NAME, INDICATION)
    nct_ids = db.load_tracking_list(DEPT_NAME, INDICATION)

    print("\n[Step 0] Baseline sync...")
    base_organized = sync_baseline(nct_ids, drug_compiled)

    state         = db.load_pipeline_state(DEPT_NAME, INDICATION)
    stored_etag   = state.get("etag")
    last_run_date = state.get("last_run_date")
    print(f"\n[State]  last_run={last_run_date or 'never'}  "
          f"etag={'set' if stored_etag else 'none'}")

    print("\n[Step 1] ETag Check...")
    has_new, new_etag = check_etag(stored_etag)
    if not has_new:
        print("\nNo new CT.gov batch. Nothing to do.")
        return

    from_date, to_date = date_range(last_run_date)

    tracking_db = {
        nct: str(row.get("Last Update Post Date", "") or "")
        for nct, row in base_organized.items()
        if str(row.get("Last Update Post Date", "") or "").strip()
    }
    prev_versions, prev_rows = db.load_version_cache(DEPT_NAME)

    print(f"\n[Step 2] FLOW 1 — {from_date} → {to_date}")
    modified_trials, new_candidates, unmatched = run_flow1(
        from_date, to_date, tracking_db, keywords, indication=INDICATION
    )

    rejected_ids = db.load_rejected_trials(DEPT_NAME, INDICATION)
    if rejected_ids:
        new_candidates = [t for t in new_candidates if t["NCT ID"] not in rejected_ids]
        unmatched      = [t for t in unmatched      if t["NCT ID"] not in rejected_ids]
        print(f"  Filtered {len(rejected_ids)} blocklisted NCTs")

    print(f"\n[Step 3] FLOW 2 — Field-Level Diff")
    all_diff_rows, newly_compared, redirects = run_flow2(
        nct_ids, prev_versions, prev_rows, drug_compiled, base_organized
    )

    print("\n[Step 4] Canonical ID auto-update...")
    nct_ids = apply_canonical_updates(redirects, nct_ids)
    if redirects:
        cleaned_set = set(nct_ids)
        for r in redirects:
            if r["requested"] not in cleaned_set:
                try:
                    db.delete_from_tracking_list(r["requested"], DEPT_NAME, INDICATION)
                except Exception:
                    pass
            if not r.get("in_input"):
                db.upsert_tracking_list(DEPT_NAME, [r["canonical"]], added_by="canonical", indication=INDICATION)

    print("\n[Step 5] Saving to DB...")
    db.insert_new_candidates(DEPT_NAME, to_date, new_candidates, indication=INDICATION)
    db.upsert_unmatched(     DEPT_NAME, to_date, unmatched,      indication=INDICATION)
    db.insert_modified_log(  DEPT_NAME, to_date, modified_trials, indication=INDICATION)
    db.insert_field_changes( DEPT_NAME, to_date, newly_compared,  indication=INDICATION)
    db.upsert_version_cache( DEPT_NAME, all_diff_rows)
    if redirects:
        db.insert_canonical_changes(DEPT_NAME, to_date, redirects)

    print("\n[Step 6c] Tagging Primary Drug...")
    tag_all_organized_drug(DEPT_NAME, drug_terms)

    _duration = (datetime.now() - run_start).total_seconds()

    print("\n[Step 7] Sending alert...")
    send_alert(
        run_date        = to_date,
        new_candidates  = new_candidates,
        unmatched       = unmatched,
        modified_trials = modified_trials,
        newly_compared  = newly_compared,
        redirects       = redirects,
    )

    print("\n[Step 8] Saving state...")
    db.save_pipeline_state(DEPT_NAME, new_etag, to_date, indication=INDICATION)
    db.insert_run_history(
        dept            = DEPT_NAME,
        run_date        = to_date,
        new_candidates  = len(new_candidates),
        unmatched       = len(unmatched),
        modified_trials = len(modified_trials),
        field_diffs     = len(newly_compared),
        canonical_fixes = len(redirects),
        duration_s      = _duration,
        output_file     = None,
        indication      = INDICATION,
    )

    print("\n" + "=" * 60)
    print(f"RUN COMPLETE  [{DEPT_NAME}]  ({_duration:.0f}s)")
    print("=" * 60)
    print(f"  New Candidates  : {len(new_candidates)}")
    print(f"  Unmatched       : {len(unmatched)}")
    print(f"  Modified        : {len(modified_trials)}")
    print(f"  Field Diffs     : {len(newly_compared)} of {len(nct_ids)} tracked")
    print(f"  Canonical Fixes : {len(redirects)}")


def main():
    parser = argparse.ArgumentParser(
        description="NCT Combined Tracking Pipeline",
        epilog=(
            "Examples:\n"
            "  python combined_pipeline.py ADC                                    # unified (default)\n"
            "  python combined_pipeline.py ADC --indication 'Ovarian Cancer'  # single pipeline\n"
            "  python combined_pipeline.py 'Liver Diseases'                   # unified Liver Diseases"
        ),
    )
    parser.add_argument("dept", help='Department name (e.g. "ADC") or full path')
    parser.add_argument(
        "--indication", default=None,
        help="Run only this one indication (for debugging). "
             "Omit to run asset + all indications in one CT.gov fetch.",
    )
    args = parser.parse_args()

    global INDICATION
    INDICATION = args.indication.strip() if args.indication is not None else None

    print("\n" + "=" * 66)
    print("NCT Combined Tracking Pipeline")
    print("=" * 66)

    print("\n[Dept]")
    init_dept(args.dept)
    run_start = datetime.now()

    if INDICATION is not None:
        # Single-pipeline debug mode: --indication "Ovarian Cancer" or --indication ""
        _run_single(run_start)
    else:
        # Default production mode: one CT.gov fetch, all pipelines classified in parallel
        INDICATION = ""
        _run_unified(run_start)


if __name__ == "__main__":
    main()
